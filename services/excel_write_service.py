"""Safe Stage 3 write-back from the website into the external IB List.

Only one existing AED and four clearly mapped fields are writable in this
stage. The official workbook is never edited in place: changes are applied to
a temporary copy, validated twice, backed up, atomically replaced, and then
read back through the Stage 2 Excel -> CSV synchronisation service.
"""

from __future__ import annotations

import csv
import json
import os
import shutil
import uuid
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator, Mapping

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from config import (
    AED_CACHE_FILE,
    CACHE_BACKUP_DIR,
    EXCEL_BACKUP_DIR,
    EXCEL_DATA_START_ROW,
    EXCEL_FILE,
    EXCEL_HEADER_ROW,
    EXCEL_SHEET,
    EXCEL_WRITE_HISTORY_FILE,
    EXCEL_WRITE_LOCK_FILE,
    MAX_CACHE_BACKUPS,
    MAX_EXCEL_BACKUPS,
    PRESERVE_CACHE_ONLY_UNITS,
    STAGING_SHEET_NAME,
    SYNC_LOCK_FILE,
    SYNC_STATE_FILE,
    TEMP_DIR,
)
from services.column_mapping import clean_excel_header, normalise_header
from services.aed_field_schema import APP_TO_EXCEL_COLUMNS, DATE_FIELDS
from services.excel_sync_service import SyncResult, sync_excel_to_cache
from utils.text_utils import clean_text


READ_ONLY_FIELDS = {
    "Serial Number",
    "Location",
    "Latitude",
    "Longitude",
    "OneMap Address",
    "Geocoding Status",
    "Excel Sync Status",
    "Patrol Schedule",
    "PM Schedule (H1)",
    "PM Schedule (H2)",
}

POSTAL_CODE_FIELDS = {"Postal Code"}
LIFT_LOBBY_FIELDS = {"Lift Lobby"}

WRITE_HISTORY_COLUMNS = [
    "Timestamp",
    "User",
    "Source Page",
    "Serial Number",
    "Field",
    "Old Value",
    "New Value",
    "Result",
    "Message",
]


@dataclass(frozen=True)
class ExcelWriteResult:
    status: str
    message: str
    serial_number: str = ""
    changed_fields: tuple[str, ...] = field(default_factory=tuple)
    backup_file: str = ""
    cache_sync_status: str = ""
    warnings: tuple[str, ...] = field(default_factory=tuple)

    @property
    def success(self) -> bool:
        return self.status == "updated"

    @property
    def excel_updated(self) -> bool:
        return self.status in {"updated", "partial"}


@dataclass(frozen=True)
class _WorkbookBaseline:
    sheet_names: tuple[str, ...]
    source_max_row: int
    source_max_column: int
    merged_ranges: tuple[str, ...]
    header_rows: tuple[tuple[Any, ...], ...]
    serial_rows: tuple[tuple[int, str], ...]
    continuation_rows: tuple[tuple[int, str], ...]


def _now_text() -> str:
    return datetime.now().astimezone().strftime("%d-%m-%Y %H:%M:%S")


def _clean_header(value: Any) -> str:
    return clean_excel_header(value or "")


def _header_map(worksheet: Any, header_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    duplicates: set[str] = set()

    for column_number in range(1, worksheet.max_column + 1):
        header = _clean_header(
            worksheet.cell(row=header_row, column=column_number).value
        )
        if not header:
            continue
        key = normalise_header(header)
        if key in result:
            duplicates.add(header)
        result[key] = column_number

    if duplicates:
        raise ValueError(
            "Duplicate Excel header values were found: "
            + ", ".join(sorted(duplicates, key=str.casefold))
        )
    return result


def _required_column(
    header_lookup: Mapping[str, int],
    excel_header: str,
) -> int:
    column_number = header_lookup.get(normalise_header(excel_header))
    if column_number is None:
        raise ValueError(f"The Excel sheet is missing required column: {excel_header}")
    return int(column_number)


def _serial_rows(
    worksheet: Any,
    *,
    serial_column: int,
    data_start_row: int,
) -> list[tuple[int, str]]:
    rows: list[tuple[int, str]] = []
    for row_number in range(data_start_row, worksheet.max_row + 1):
        value = clean_text(worksheet.cell(row=row_number, column=serial_column).value)
        if value:
            rows.append((row_number, value))
    return rows


def _find_serial_row(
    worksheet: Any,
    *,
    serial_column: int,
    data_start_row: int,
    serial_number: str,
) -> int:
    wanted = clean_text(serial_number).casefold()
    matches = [
        row_number
        for row_number, value in _serial_rows(
            worksheet,
            serial_column=serial_column,
            data_start_row=data_start_row,
        )
        if value.casefold() == wanted
    ]

    if not matches:
        raise ValueError(f"Unit not found in Excel: {serial_number}")
    if len(matches) > 1:
        raise ValueError(f"Duplicate Serial Number found in Excel: {serial_number}")
    return matches[0]


def _continuation_rows_after(
    worksheet: Any,
    *,
    row_number: int,
    serial_column: int,
) -> list[int]:
    """Return blank-serial rows belonging to the selected AED until the next unit."""
    rows: list[int] = []
    for candidate in range(row_number + 1, worksheet.max_row + 1):
        if clean_text(worksheet.cell(candidate, serial_column).value):
            break
        rows.append(candidate)
    return rows


def _read_combined_remarks(
    worksheet: Any,
    *,
    row_number: int,
    serial_column: int,
    remarks_column: int,
) -> str:
    parts = [clean_text(worksheet.cell(row_number, remarks_column).value)]
    for continuation_row in _continuation_rows_after(
        worksheet, row_number=row_number, serial_column=serial_column
    ):
        parts.append(clean_text(worksheet.cell(continuation_row, remarks_column).value))
    return " ".join(part for part in parts if part).strip()


def _write_combined_remarks(
    worksheet: Any,
    *,
    row_number: int,
    serial_column: int,
    remarks_column: int,
    value: Any,
) -> list[int]:
    """Consolidate all Remarks into the main unit row and clear old continuation text."""
    worksheet.cell(row_number, remarks_column).value = clean_text(value) or None
    continuation_rows = _continuation_rows_after(
        worksheet, row_number=row_number, serial_column=serial_column
    )
    for continuation_row in continuation_rows:
        worksheet.cell(continuation_row, remarks_column).value = None
    return continuation_rows


def _normalise_postal_code(value: Any) -> str:
    raw = clean_text(value)
    if raw.casefold() in {"", "n/a", "na", "nil", "none", "nan", "-"}:
        return ""
    text = raw.replace(" ", "")
    if not text.isdigit():
        raise ValueError("Postal Code must contain digits only.")
    if len(text) > 6:
        raise ValueError("Postal Code must contain at most six digits.")
    return text.zfill(6)


def _normalise_date(value: Any) -> date | None:
    if value is None or clean_text(value).casefold() in {"", "n/a", "na", "nil", "none", "nan", "-"}:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean_text(value)
    for date_format in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError("The supplied date value must be a real date.")


def _normalise_text(value: Any) -> str:
    text = clean_text(value)
    if text.casefold() in {"", "n/a", "na", "nil", "none", "nan", "-"}:
        return ""
    return text


def _normalise_lift_lobby(value: Any) -> str:
    text = _normalise_text(value)
    if not text:
        return ""
    lowered = text.casefold()
    marker = "lift lobby"
    position = lowered.rfind(marker)
    if position >= 0:
        return text[position + len(marker):].strip()
    return text


def _display_value(app_field: str, value: Any) -> str:
    if value is None:
        return ""
    if app_field in DATE_FIELDS:
        parsed = _normalise_date(value)
        return "" if parsed is None else parsed.strftime("%d-%m-%Y")
    if app_field in POSTAL_CODE_FIELDS:
        return _normalise_postal_code(value)
    if app_field in LIFT_LOBBY_FIELDS:
        return _normalise_lift_lobby(value)
    return _normalise_text(value)


def _normalise_changes(changes: Mapping[str, Any]) -> dict[str, Any]:
    if not isinstance(changes, Mapping):
        raise ValueError("Changes must be supplied as a field-to-value mapping.")

    unsupported: list[str] = []
    read_only: list[str] = []
    normalised: dict[str, Any] = {}

    for raw_field, raw_value in changes.items():
        field_name = clean_text(raw_field)
        if field_name in READ_ONLY_FIELDS:
            read_only.append(field_name)
            continue
        if field_name not in APP_TO_EXCEL_COLUMNS:
            unsupported.append(field_name or "<blank field>")
            continue

        if field_name in DATE_FIELDS:
            normalised[field_name] = _normalise_date(raw_value)
        elif field_name in POSTAL_CODE_FIELDS:
            normalised[field_name] = _normalise_postal_code(raw_value)
        elif field_name in LIFT_LOBBY_FIELDS:
            normalised[field_name] = _normalise_lift_lobby(raw_value)
        else:
            normalised[field_name] = _normalise_text(raw_value)

    if read_only:
        raise ValueError(
            "These fields are read-only: "
            + ", ".join(sorted(set(read_only), key=str.casefold))
        )
    if unsupported:
        raise ValueError(
            "Unsupported write-back field(s): "
            + ", ".join(sorted(set(unsupported), key=str.casefold))
        )
    return normalised


def _read_app_value(cell: Any, app_field: str) -> str:
    return _display_value(app_field, cell.value)


def _write_cell(cell: Any, app_field: str, value: Any) -> None:
    if isinstance(cell, MergedCell):
        raise ValueError(
            f"The target Excel cell for {app_field} is not the top-left cell "
            "of its merged range."
        )

    if app_field in DATE_FIELDS:
        cell.value = value
        cell.number_format = "dd/mm/yyyy"
        return

    if app_field in POSTAL_CODE_FIELDS:
        cell.value = value or None
        cell.number_format = "@"
        return

    cell.value = value or None


def _capture_baseline(
    workbook: Any,
    *,
    sheet_name: str,
    header_row: int,
    data_start_row: int,
) -> _WorkbookBaseline:
    worksheet = workbook[sheet_name]
    lookup = _header_map(worksheet, header_row)
    serial_column = _required_column(lookup, "SERIAL NUMBER")
    remarks_column = _required_column(lookup, "Remarks")

    header_rows = tuple(
        tuple(
            worksheet.cell(row=row_number, column=column_number).value
            for column_number in range(1, worksheet.max_column + 1)
        )
        for row_number in range(1, data_start_row)
    )
    serial_rows = tuple(
        _serial_rows(
            worksheet,
            serial_column=serial_column,
            data_start_row=data_start_row,
        )
    )
    continuation_rows = tuple(
        (
            row_number,
            clean_text(worksheet.cell(row=row_number, column=remarks_column).value),
        )
        for row_number in range(data_start_row, worksheet.max_row + 1)
        if not clean_text(worksheet.cell(row=row_number, column=serial_column).value)
        and clean_text(worksheet.cell(row=row_number, column=remarks_column).value)
    )

    return _WorkbookBaseline(
        sheet_names=tuple(workbook.sheetnames),
        source_max_row=worksheet.max_row,
        source_max_column=worksheet.max_column,
        merged_ranges=tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)),
        header_rows=header_rows,
        serial_rows=serial_rows,
        continuation_rows=continuation_rows,
    )


def _validate_structure(
    workbook: Any,
    *,
    baseline: _WorkbookBaseline,
    sheet_name: str,
    header_row: int,
    data_start_row: int,
    staging_expected: bool,
    allowed_continuation_rows: set[int] | None = None,
) -> None:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' disappeared during validation.")
    if staging_expected and STAGING_SHEET_NAME not in workbook.sheetnames:
        raise ValueError("The staging worksheet is missing during validation.")
    if not staging_expected and STAGING_SHEET_NAME in workbook.sheetnames:
        raise ValueError("The staging worksheet was not removed before replacement.")

    expected_sheet_names = list(baseline.sheet_names)
    if staging_expected:
        expected_sheet_names.append(STAGING_SHEET_NAME)
    if tuple(workbook.sheetnames) != tuple(expected_sheet_names):
        raise ValueError("The workbook worksheet list changed unexpectedly.")

    worksheet = workbook[sheet_name]
    if worksheet.max_row != baseline.source_max_row:
        raise ValueError("The source worksheet row count changed unexpectedly.")
    if worksheet.max_column != baseline.source_max_column:
        raise ValueError("The source worksheet column count changed unexpectedly.")
    if tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)) != (
        baseline.merged_ranges
    ):
        raise ValueError("Merged-cell ranges changed unexpectedly.")

    current_headers = tuple(
        tuple(
            worksheet.cell(row=row_number, column=column_number).value
            for column_number in range(1, worksheet.max_column + 1)
        )
        for row_number in range(1, data_start_row)
    )
    if current_headers != baseline.header_rows:
        raise ValueError("The first two Excel header rows changed unexpectedly.")

    lookup = _header_map(worksheet, header_row)
    serial_column = _required_column(lookup, "SERIAL NUMBER")
    remarks_column = _required_column(lookup, "Remarks")
    if tuple(
        _serial_rows(
            worksheet,
            serial_column=serial_column,
            data_start_row=data_start_row,
        )
    ) != baseline.serial_rows:
        raise ValueError("The Serial Number rows changed unexpectedly.")

    allowed = set(allowed_continuation_rows or set())
    baseline_continuations = dict(baseline.continuation_rows)
    for row_number, old_value in baseline_continuations.items():
        if clean_text(worksheet.cell(row=row_number, column=serial_column).value):
            raise ValueError("A Remarks continuation row gained a Serial Number unexpectedly.")
        current_value = clean_text(worksheet.cell(row=row_number, column=remarks_column).value)
        if row_number not in allowed and current_value != old_value:
            raise ValueError("Remarks continuation rows changed unexpectedly.")

    for row_number in range(data_start_row, worksheet.max_row + 1):
        if row_number in baseline_continuations or row_number in allowed:
            continue
        if (
            not clean_text(worksheet.cell(row=row_number, column=serial_column).value)
            and clean_text(worksheet.cell(row=row_number, column=remarks_column).value)
        ):
            raise ValueError("A new Remarks continuation row appeared unexpectedly.")


def _validate_target_values(
    workbook: Any,
    *,
    sheet_name: str,
    header_row: int,
    data_start_row: int,
    serial_number: str,
    expected_changes: Mapping[str, Any],
) -> None:
    worksheet = workbook[sheet_name]
    lookup = _header_map(worksheet, header_row)
    serial_column = _required_column(lookup, "SERIAL NUMBER")
    row_number = _find_serial_row(
        worksheet,
        serial_column=serial_column,
        data_start_row=data_start_row,
        serial_number=serial_number,
    )

    for app_field, expected_value in expected_changes.items():
        excel_header = APP_TO_EXCEL_COLUMNS[app_field]
        column_number = _required_column(lookup, excel_header)
        if app_field == "Remarks":
            actual = _read_combined_remarks(
                worksheet,
                row_number=row_number,
                serial_column=serial_column,
                remarks_column=column_number,
            )
        else:
            actual = _read_app_value(
                worksheet.cell(row=row_number, column=column_number),
                app_field,
            )
        expected = _display_value(app_field, expected_value)
        if actual != expected:
            raise ValueError(
                f"Temporary workbook validation failed for {app_field}: "
                f"expected '{expected}', found '{actual}'."
            )


def _create_staging_sheet(
    workbook: Any,
    *,
    serial_number: str,
    row_number: int,
    changes: Mapping[str, Any],
    old_values: Mapping[str, str],
    page_original_values: Mapping[str, str],
    user: str,
    source_page: str,
) -> None:
    if STAGING_SHEET_NAME in workbook.sheetnames:
        raise ValueError(
            f"The workbook already contains reserved sheet {STAGING_SHEET_NAME}."
        )

    staging = workbook.create_sheet(STAGING_SHEET_NAME)
    staging.append(
        [
            "Timestamp",
            "User",
            "Source Page",
            "Serial Number",
            "App Field",
            "Excel Column",
            "Excel Row",
            "Page Original Value",
            "Current Excel Value",
            "New Value",
        ]
    )
    timestamp = _now_text()
    for app_field, new_value in changes.items():
        staging.append(
            [
                timestamp,
                user,
                source_page,
                serial_number,
                app_field,
                APP_TO_EXCEL_COLUMNS[app_field],
                row_number,
                page_original_values.get(app_field, ""),
                old_values.get(app_field, ""),
                _display_value(app_field, new_value),
            ]
        )
    staging.freeze_panes = "A2"
    staging.sheet_state = "hidden"


def _make_temporary_workbook_path(source: Path) -> Path:
    return source.parent / (
        f".{source.stem}_write_{uuid.uuid4().hex[:10]}{source.suffix}"
    )


def _copy_stable_workbook(source: Path, destination: Path) -> None:
    before = source.stat()
    shutil.copy2(source, destination)
    after = source.stat()
    if (
        before.st_mtime_ns != after.st_mtime_ns
        or before.st_size != after.st_size
    ):
        destination.unlink(missing_ok=True)
        raise RuntimeError(
            "The Excel workbook changed while it was being copied. "
            "Refresh the page and try again."
        )


def _create_excel_backup(
    source: Path,
    backup_dir: Path,
    max_backups: int,
) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / (
        f"{source.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{source.suffix}"
    )
    shutil.copy2(source, backup)

    backups = sorted(
        backup_dir.glob(f"{source.stem}_*{source.suffix}"),
        key=lambda item: item.stat().st_mtime_ns,
        reverse=True,
    )
    for old_backup in backups[max(1, int(max_backups)) :]:
        old_backup.unlink(missing_ok=True)
    return backup


@contextmanager
def _acquire_write_lock(
    lock_file: Path,
    *,
    user: str,
    serial_number: str,
) -> Iterator[None]:
    lock_file.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "user": user,
        "serial_number": serial_number,
        "started_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "operation": "update_existing_unit",
    }

    try:
        descriptor = os.open(
            lock_file,
            os.O_CREAT | os.O_EXCL | os.O_WRONLY,
            0o600,
        )
    except FileExistsError as error:
        detail = ""
        try:
            existing = json.loads(lock_file.read_text(encoding="utf-8"))
            owner = clean_text(existing.get("user", ""))
            started = clean_text(existing.get("started_at", ""))
            if owner or started:
                detail = f" Current lock: {owner or 'unknown user'}, {started}."
        except (OSError, json.JSONDecodeError, AttributeError):
            pass
        raise RuntimeError(
            "Another Excel write is already in progress. Try again after it finishes."
            + detail
        ) from error

    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, ensure_ascii=False)
        yield
    finally:
        lock_file.unlink(missing_ok=True)


def _append_history(
    history_file: Path,
    *,
    serial_number: str,
    changes: Mapping[str, Any],
    old_values: Mapping[str, str],
    user: str,
    source_page: str,
    result: str,
    message: str,
) -> None:
    history_file.parent.mkdir(parents=True, exist_ok=True)
    write_header = not history_file.exists() or history_file.stat().st_size == 0
    timestamp = _now_text()
    rows = []

    if changes:
        for field_name, new_value in changes.items():
            rows.append(
                {
                    "Timestamp": timestamp,
                    "User": user,
                    "Source Page": source_page,
                    "Serial Number": serial_number,
                    "Field": field_name,
                    "Old Value": old_values.get(field_name, ""),
                    "New Value": _display_value(field_name, new_value),
                    "Result": result,
                    "Message": message,
                }
            )
    else:
        rows.append(
            {
                "Timestamp": timestamp,
                "User": user,
                "Source Page": source_page,
                "Serial Number": serial_number,
                "Field": "",
                "Old Value": "",
                "New Value": "",
                "Result": result,
                "Message": message,
            }
        )

    with history_file.open("a", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=WRITE_HISTORY_COLUMNS)
        if write_header:
            writer.writeheader()
        writer.writerows(rows)


def load_excel_write_history(
    history_file: str | Path = EXCEL_WRITE_HISTORY_FILE,
) -> list[dict[str, str]]:
    path = Path(history_file)
    if not path.exists() or path.stat().st_size == 0:
        return []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def update_existing_unit_in_excel(
    *,
    serial_number: str,
    changes: Mapping[str, Any],
    original_values: Mapping[str, Any] | None,
    user: str,
    source_page: str,
    excel_file: str | Path = EXCEL_FILE,
    excel_sheet: str = EXCEL_SHEET,
    write_lock_file: str | Path = EXCEL_WRITE_LOCK_FILE,
    excel_backup_dir: str | Path = EXCEL_BACKUP_DIR,
    write_history_file: str | Path = EXCEL_WRITE_HISTORY_FILE,
    cache_file: str | Path = AED_CACHE_FILE,
    sync_state_file: str | Path = SYNC_STATE_FILE,
    sync_lock_file: str | Path = SYNC_LOCK_FILE,
    cache_backup_dir: str | Path = CACHE_BACKUP_DIR,
    temp_dir: str | Path = TEMP_DIR,
    header_row: int = EXCEL_HEADER_ROW,
    data_start_row: int = EXCEL_DATA_START_ROW,
    max_excel_backups: int = MAX_EXCEL_BACKUPS,
    max_cache_backups: int = MAX_CACHE_BACKUPS,
    preserve_cache_only_units: bool = PRESERVE_CACHE_ONLY_UNITS,
) -> ExcelWriteResult:
    """Safely update one existing AED in the IB List and refresh the CSV mirror."""

    serial = clean_text(serial_number)
    changed_by = clean_text(user)
    page = clean_text(source_page) or "Unknown"
    if not serial:
        return ExcelWriteResult(status="failed", message="Serial Number is required.")
    if not changed_by:
        return ExcelWriteResult(
            status="failed",
            message="Changed By is required so the update can be audited.",
            serial_number=serial,
        )

    try:
        normalised_changes = _normalise_changes(changes)
    except ValueError as error:
        return ExcelWriteResult(
            status="failed",
            message=str(error),
            serial_number=serial,
        )

    if not normalised_changes:
        return ExcelWriteResult(
            status="no_changes",
            message="No changes were detected.",
            serial_number=serial,
        )

    source = Path(excel_file)
    lock_path = Path(write_lock_file)
    backup_dir = Path(excel_backup_dir)
    history_path = Path(write_history_file)
    cache_path = Path(cache_file)
    sync_state_path = Path(sync_state_file)
    sync_lock_path = Path(sync_lock_file)
    cache_backup_path = Path(cache_backup_dir)
    temp_directory = Path(temp_dir)

    for directory in {
        source.parent,
        lock_path.parent,
        backup_dir,
        history_path.parent,
        cache_path.parent,
        sync_state_path.parent,
        sync_lock_path.parent,
        cache_backup_path,
        temp_directory,
    }:
        directory.mkdir(parents=True, exist_ok=True)

    if not source.exists():
        return ExcelWriteResult(
            status="failed",
            message=f"Excel workbook not found: {source}",
            serial_number=serial,
        )

    temporary_workbook: Path | None = None
    backup_file: Path | None = None
    actual_changes: dict[str, Any] = {}
    old_values: dict[str, str] = {}
    page_original_values = {
        field_name: _display_value(field_name, value)
        for field_name, value in (original_values or {}).items()
        if field_name in APP_TO_EXCEL_COLUMNS
    }
    excel_replaced = False

    try:
        with _acquire_write_lock(
            lock_path,
            user=changed_by,
            serial_number=serial,
        ):
            temporary_workbook = _make_temporary_workbook_path(source)
            _copy_stable_workbook(source, temporary_workbook)

            workbook = load_workbook(temporary_workbook, data_only=False)
            if excel_sheet not in workbook.sheetnames:
                raise ValueError(
                    f"Worksheet '{excel_sheet}' was not found. Available sheets: "
                    + ", ".join(workbook.sheetnames)
                )

            baseline = _capture_baseline(
                workbook,
                sheet_name=excel_sheet,
                header_row=header_row,
                data_start_row=data_start_row,
            )
            worksheet = workbook[excel_sheet]
            headers = _header_map(worksheet, header_row)
            serial_column = _required_column(headers, "SERIAL NUMBER")
            row_number = _find_serial_row(
                worksheet,
                serial_column=serial_column,
                data_start_row=data_start_row,
                serial_number=serial,
            )

            column_numbers: dict[str, int] = {}
            for app_field, new_value in normalised_changes.items():
                excel_header = APP_TO_EXCEL_COLUMNS[app_field]
                column_number = _required_column(headers, excel_header)
                column_numbers[app_field] = column_number
                cell = worksheet.cell(row=row_number, column=column_number)
                old_value = _read_app_value(cell, app_field)
                new_display = _display_value(app_field, new_value)
                old_values[app_field] = old_value
                if old_value != new_display:
                    actual_changes[app_field] = new_value

            if not actual_changes:
                workbook.close()
                temporary_workbook.unlink(missing_ok=True)
                return ExcelWriteResult(
                    status="no_changes",
                    message="No changes were detected.",
                    serial_number=serial,
                )

            _create_staging_sheet(
                workbook,
                serial_number=serial,
                row_number=row_number,
                changes=actual_changes,
                old_values=old_values,
                page_original_values=page_original_values,
                user=changed_by,
                source_page=page,
            )

            for app_field, new_value in actual_changes.items():
                cell = worksheet.cell(
                    row=row_number,
                    column=column_numbers[app_field],
                )
                _write_cell(cell, app_field, new_value)

            workbook.save(temporary_workbook)
            workbook.close()

            first_check = load_workbook(temporary_workbook, data_only=False)
            _validate_structure(
                first_check,
                baseline=baseline,
                sheet_name=excel_sheet,
                header_row=header_row,
                data_start_row=data_start_row,
                staging_expected=True,
            )
            _validate_target_values(
                first_check,
                sheet_name=excel_sheet,
                header_row=header_row,
                data_start_row=data_start_row,
                serial_number=serial,
                expected_changes=actual_changes,
            )
            del first_check[STAGING_SHEET_NAME]
            first_check.save(temporary_workbook)
            first_check.close()

            second_check = load_workbook(temporary_workbook, data_only=False)
            _validate_structure(
                second_check,
                baseline=baseline,
                sheet_name=excel_sheet,
                header_row=header_row,
                data_start_row=data_start_row,
                staging_expected=False,
            )
            _validate_target_values(
                second_check,
                sheet_name=excel_sheet,
                header_row=header_row,
                data_start_row=data_start_row,
                serial_number=serial,
                expected_changes=actual_changes,
            )
            second_check.close()

            backup_file = _create_excel_backup(
                source,
                backup_dir,
                max_excel_backups,
            )

            try:
                os.replace(temporary_workbook, source)
            except PermissionError as error:
                raise PermissionError(
                    "The Excel workbook is currently open or locked. "
                    "Close the workbook and try again."
                ) from error
            temporary_workbook = None
            excel_replaced = True

            try:
                sync_result: SyncResult = sync_excel_to_cache(
                force=True,
                excel_file=source,
                excel_sheet=excel_sheet,
                cache_file=cache_path,
                state_file=sync_state_path,
                temp_dir=temp_directory,
                lock_file=sync_lock_path,
                backup_dir=cache_backup_path,
                preserve_cache_only_units=preserve_cache_only_units,
                    max_backups=max_cache_backups,
                )
            except Exception as error:
                message = (
                    "Excel was updated successfully, but the website cache could "
                    f"not be refreshed. Use Refresh AED Data. {error}"
                )
                try:
                    _append_history(
                        history_path,
                        serial_number=serial,
                        changes=actual_changes,
                        old_values=old_values,
                        user=changed_by,
                        source_page=page,
                        result="Excel Updated; Cache Refresh Failed",
                        message=message,
                    )
                except OSError:
                    pass
                return ExcelWriteResult(
                    status="partial",
                    message=message,
                    serial_number=serial,
                    changed_fields=tuple(actual_changes),
                    backup_file=str(backup_file or ""),
                    cache_sync_status="exception",
                )

            if not sync_result.success:
                message = (
                    "Excel was updated successfully, but the website cache could "
                    "not be refreshed. Use Refresh AED Data. "
                    + sync_result.message
                )
                try:
                    _append_history(
                        history_path,
                        serial_number=serial,
                        changes=actual_changes,
                        old_values=old_values,
                        user=changed_by,
                        source_page=page,
                        result="Excel Updated; Cache Refresh Failed",
                        message=message,
                    )
                except OSError:
                    pass
                return ExcelWriteResult(
                    status="partial",
                    message=message,
                    serial_number=serial,
                    changed_fields=tuple(actual_changes),
                    backup_file=str(backup_file or ""),
                    cache_sync_status=sync_result.status,
                    warnings=tuple(sync_result.warnings),
                )

            message = (
                f"Updated {serial} in Excel and refreshed the website data "
                f"({len(actual_changes)} field change(s))."
            )
            history_warning = ""
            try:
                _append_history(
                    history_path,
                    serial_number=serial,
                    changes=actual_changes,
                    old_values=old_values,
                    user=changed_by,
                    source_page=page,
                    result="Success",
                    message=message,
                )
            except OSError as error:
                history_warning = f"The write succeeded, but audit history failed: {error}"

            warnings = list(sync_result.warnings)
            if history_warning:
                warnings.append(history_warning)
            return ExcelWriteResult(
                status="updated",
                message=message,
                serial_number=serial,
                changed_fields=tuple(actual_changes),
                backup_file=str(backup_file or ""),
                cache_sync_status=sync_result.status,
                warnings=tuple(warnings),
            )

    except (OSError, ValueError, RuntimeError) as error:
        message = str(error)
        history_result = "Excel Updated; Post-write Step Failed" if excel_replaced else "Failed"
        if actual_changes:
            try:
                _append_history(
                    history_path,
                    serial_number=serial,
                    changes=actual_changes,
                    old_values=old_values,
                    user=changed_by,
                    source_page=page,
                    result=history_result,
                    message=message,
                )
            except OSError:
                pass
        return ExcelWriteResult(
            status="partial" if excel_replaced else "failed",
            message=(
                "Excel was updated, but a later step failed: " + message
                if excel_replaced
                else message
            ),
            serial_number=serial,
            changed_fields=tuple(actual_changes),
            backup_file=str(backup_file or ""),
        )
    finally:
        if temporary_workbook is not None:
            temporary_workbook.unlink(missing_ok=True)
