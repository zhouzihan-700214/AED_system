"""Stage 4 transaction coordinator for safe multi-user Excel operations."""
from __future__ import annotations

import csv
import json
import os
import shutil
import uuid
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook

from config import (
    ACTIVE_TRANSACTION_FILE,
    AED_CACHE_FILE,
    AED_LIFECYCLE_FILE,
    CACHE_BACKUP_DIR,
    EXCEL_BACKUP_DIR,
    EXCEL_DATA_START_ROW,
    EXCEL_FILE,
    EXCEL_HEADER_ROW,
    EXCEL_OPERATION_LOCK_FILE,
    EXCEL_SHEET,
    MAX_CACHE_BACKUPS,
    MAX_EXCEL_BACKUPS,
    PRESERVE_CACHE_ONLY_UNITS,
    STAGING_SHEET_NAME,
    SYNC_STATE_FILE,
    TEMP_DIR,
)
from services.audit_service import record_conflicts, record_field_audit, record_transaction
from services.onedrive_excel_service import is_cloud_onedrive_enabled
from services.conflict_service import detect_field_conflicts
from services.excel_lock_service import operation_lock
from services.excel_sync_service import SyncResult, sync_excel_to_cache
from services.excel_write_service import (
    APP_TO_EXCEL_COLUMNS,
    DATE_FIELDS,
    _capture_baseline,
    _copy_stable_workbook,
    _continuation_rows_after,
    _create_excel_backup,
    _display_value,
    _find_serial_row,
    _header_map,
    _make_temporary_workbook_path,
    _normalise_changes,
    _normalise_lift_lobby,
    _read_app_value,
    _read_combined_remarks,
    _required_column,
    _serial_rows,
    _validate_structure,
    _validate_target_values,
    _write_cell,
    _write_combined_remarks,
)
from utils.text_utils import clean_text


ADD_TO_EXCEL_COLUMNS = dict(APP_TO_EXCEL_COLUMNS)
DATE_ADD_FIELDS = set(DATE_FIELDS)
LIFECYCLE_COLUMNS = [
    "Timestamp", "Operation ID", "User", "Session ID", "Serial Number",
    "Status", "Reason", "Source Page",
]


@dataclass(frozen=True)
class OperationResult:
    status: str
    message: str
    operation_id: str = ""
    serial_number: str = ""
    serial_numbers: tuple[str, ...] = field(default_factory=tuple)
    changed_fields: tuple[str, ...] = field(default_factory=tuple)
    backup_file: str = ""
    cache_sync_status: str = ""
    warnings: tuple[str, ...] = field(default_factory=tuple)
    conflicts: dict[str, Any] = field(default_factory=dict)
    already_applied: tuple[str, ...] = field(default_factory=tuple)

    @property
    def success(self) -> bool:
        return self.status in {"updated", "added", "deactivated"}

    @property
    def excel_updated(self) -> bool:
        return self.status in {"updated", "added", "partial"}


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _atomic_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    temp.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    os.replace(temp, path)


def _journal(operation: dict[str, Any], step: str, status: str | None = None, **updates: Any) -> None:
    operation.update(updates)
    operation["last_completed_step"] = step
    if status:
        operation["status"] = status
    _atomic_json(Path(ACTIVE_TRANSACTION_FILE), operation)


def _finish_journal(operation: dict[str, Any], result: str, message: str, *, clear_active: bool = True) -> None:
    record_transaction(
        **{
            "Operation ID": operation["operation_id"],
            "Operation Type": operation["operation_type"],
            "User": operation.get("user", ""),
            "Session ID": operation.get("session_id", ""),
            "Serial Numbers": "; ".join(operation.get("serial_numbers", [])),
            "Source Page": operation.get("source_page", ""),
            "Started At": operation.get("started_at", ""),
            "Result": result,
            "Final Step": operation.get("last_completed_step", ""),
            "Message": message,
        }
    )
    if clear_active:
        Path(ACTIVE_TRANSACTION_FILE).unlink(missing_ok=True)


def _new_operation(operation_type: str, user: str, session_id: str, serials: Sequence[str], source_page: str) -> dict[str, Any]:
    operation_id = str(uuid.uuid4())
    return {
        "operation_id": operation_id,
        "operation_type": operation_type,
        "status": "STARTED",
        "user": clean_text(user),
        "session_id": clean_text(session_id),
        "serial_numbers": [clean_text(item) for item in serials],
        "source_page": clean_text(source_page),
        "source_file": str(EXCEL_FILE),
        "temp_file": None,
        "backup_file": None,
        "started_at": _now_iso(),
        "last_completed_step": "STARTED",
    }


def _sync_without_relocking() -> SyncResult:
    return sync_excel_to_cache(
        force=True,
        excel_file=EXCEL_FILE,
        excel_sheet=EXCEL_SHEET,
        cache_file=AED_CACHE_FILE,
        state_file=SYNC_STATE_FILE,
        temp_dir=TEMP_DIR,
        lock_file=EXCEL_OPERATION_LOCK_FILE,
        backup_dir=CACHE_BACKUP_DIR,
        preserve_cache_only_units=(False if is_cloud_onedrive_enabled() else PRESERVE_CACHE_ONLY_UNITS),
        max_backups=MAX_CACHE_BACKUPS,
        acquire_lock=False,
    )


def _audit_rows(
    operation: Mapping[str, Any], serial: str, desired: Mapping[str, Any], original: Mapping[str, Any],
    current: Mapping[str, Any], result: str, message: str,
) -> list[dict[str, Any]]:
    return [
        {
            "Operation ID": operation["operation_id"],
            "User": operation.get("user", ""),
            "Session ID": operation.get("session_id", ""),
            "Operation Type": operation.get("operation_type", ""),
            "Source Page": operation.get("source_page", ""),
            "Serial Number": serial,
            "Field": field_name,
            "Original Value": original.get(field_name, ""),
            "Current Value": current.get(field_name, ""),
            "Desired Value": _display_value(field_name, desired_value),
            "Result": result,
            "Message": message,
        }
        for field_name, desired_value in desired.items()
    ]


def execute_unit_update(
    *, serial_number: str, desired_values: Mapping[str, Any], original_values: Mapping[str, Any],
    user: str, session_id: str, source_page: str,
) -> OperationResult:
    serial = clean_text(serial_number)
    changed_by = clean_text(user)
    session = clean_text(session_id)
    if not serial:
        return OperationResult("failed", "Serial Number is required.")
    if not changed_by:
        return OperationResult("failed", "Select an audit user before saving.", serial_number=serial)
    try:
        normalised = _normalise_changes(desired_values)
    except ValueError as error:
        return OperationResult("failed", str(error), serial_number=serial)
    if not normalised:
        return OperationResult("no_changes", "No changes were detected.", serial_number=serial)

    operation = _new_operation("UPDATE_UNIT", changed_by, session, [serial], source_page)
    temp_path: Path | None = None
    backup: Path | None = None
    source_replaced = False
    try:
        with operation_lock({
            "operation_id": operation["operation_id"], "operation_type": "UPDATE_UNIT",
            "user": changed_by, "session_id": session, "serial_numbers": [serial],
        }, EXCEL_OPERATION_LOCK_FILE):
            _journal(operation, "LOCK_ACQUIRED", "LOCK_ACQUIRED")
            temp_path = _make_temporary_workbook_path(Path(EXCEL_FILE))
            operation["temp_file"] = str(temp_path)
            _copy_stable_workbook(Path(EXCEL_FILE), temp_path)
            workbook = load_workbook(temp_path, data_only=False)
            if EXCEL_SHEET not in workbook.sheetnames:
                raise ValueError(f"Worksheet '{EXCEL_SHEET}' was not found.")
            _journal(operation, "SOURCE_READ", "SOURCE_READ")

            baseline = _capture_baseline(workbook, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW)
            worksheet = workbook[EXCEL_SHEET]
            headers = _header_map(worksheet, EXCEL_HEADER_ROW)
            serial_col = _required_column(headers, "SERIAL NUMBER")
            row_number = _find_serial_row(worksheet, serial_column=serial_col, data_start_row=EXCEL_DATA_START_ROW, serial_number=serial)

            current: dict[str, str] = {}
            columns: dict[str, int] = {}
            original_display: dict[str, str] = {}
            desired_display: dict[str, str] = {}
            for field_name, desired_value in normalised.items():
                col = _required_column(headers, APP_TO_EXCEL_COLUMNS[field_name])
                columns[field_name] = col
                if field_name == "Remarks":
                    current[field_name] = _read_combined_remarks(
                        worksheet,
                        row_number=row_number,
                        serial_column=serial_col,
                        remarks_column=col,
                    )
                else:
                    current[field_name] = _read_app_value(
                        worksheet.cell(row=row_number, column=col), field_name
                    )
                original_display[field_name] = _display_value(field_name, original_values.get(field_name, ""))
                desired_display[field_name] = _display_value(field_name, desired_value)

            conflict_check = detect_field_conflicts(original_display, current, desired_display)
            _journal(operation, "CONFLICT_CHECKED", "CONFLICT_CHECKED")
            if conflict_check["conflicts"]:
                message = "This AED was changed after you opened the page. Review the conflicting fields before saving again."
                rows = _audit_rows(operation, serial, normalised, original_display, current, "Conflict", message)
                record_conflicts([row for row in rows if row["Field"] in conflict_check["conflicts"]])
                record_field_audit(rows)
                workbook.close()
                temp_path.unlink(missing_ok=True)
                temp_path = None
                _journal(operation, "CONFLICT_DETECTED", "CONFLICT_DETECTED")
                _finish_journal(operation, "Conflict", message)
                return OperationResult(
                    "conflict", message, operation_id=operation["operation_id"], serial_number=serial,
                    conflicts=conflict_check["conflicts"], already_applied=tuple(conflict_check["already_applied"]),
                )

            safe_changes = {field: normalised[field] for field in conflict_check["safe_changes"]}
            if not safe_changes:
                message = "The requested values were already present in Excel. No additional update was required."
                rows = _audit_rows(operation, serial, normalised, original_display, current, "Already Applied", message)
                record_field_audit(rows)
                workbook.close()
                temp_path.unlink(missing_ok=True)
                temp_path = None
                _finish_journal(operation, "Already Applied", message)
                return OperationResult(
                    "already_applied", message, operation_id=operation["operation_id"], serial_number=serial,
                    already_applied=tuple(conflict_check["already_applied"]),
                )

            # Build the hidden staging sheet without changing the official source sheet yet.
            staging = workbook.create_sheet(STAGING_SHEET_NAME)
            staging.sheet_state = "hidden"
            staging.append(["Operation ID", "Serial Number", "Field", "Excel Row", "Original", "Current", "Desired"])
            allowed_continuation_rows: set[int] = set()
            for field_name, value in safe_changes.items():
                staging.append([
                    operation["operation_id"], serial, field_name, row_number,
                    original_display.get(field_name, ""), current.get(field_name, ""), _display_value(field_name, value),
                ])
                if field_name == "Remarks":
                    allowed_continuation_rows.update(
                        _write_combined_remarks(
                            worksheet,
                            row_number=row_number,
                            serial_column=serial_col,
                            remarks_column=columns[field_name],
                            value=value,
                        )
                    )
                else:
                    _write_cell(
                        worksheet.cell(row=row_number, column=columns[field_name]),
                        field_name,
                        value,
                    )
            _journal(operation, "CHANGES_APPLIED", "CHANGES_APPLIED")
            workbook.save(temp_path)
            workbook.close()

            first = load_workbook(temp_path, data_only=False)
            _validate_structure(first, baseline=baseline, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW, staging_expected=True, allowed_continuation_rows=allowed_continuation_rows)
            _validate_target_values(first, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW, serial_number=serial, expected_changes=safe_changes)
            del first[STAGING_SHEET_NAME]
            first.save(temp_path)
            first.close()
            _journal(operation, "TEMP_VALIDATED", "TEMP_VALIDATED")

            second = load_workbook(temp_path, data_only=False)
            _validate_structure(second, baseline=baseline, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW, staging_expected=False, allowed_continuation_rows=allowed_continuation_rows)
            _validate_target_values(second, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW, serial_number=serial, expected_changes=safe_changes)
            second.close()

            backup = _create_excel_backup(Path(EXCEL_FILE), Path(EXCEL_BACKUP_DIR), MAX_EXCEL_BACKUPS)
            operation["backup_file"] = str(backup)
            _journal(operation, "BACKUP_CREATED", "BACKUP_CREATED")
            try:
                os.replace(temp_path, EXCEL_FILE)
            except PermissionError as error:
                raise PermissionError("The Excel workbook is open or locked. No data was changed.") from error
            temp_path = None
            source_replaced = True
            _journal(operation, "SOURCE_REPLACED", "SOURCE_REPLACED")

            sync_result = _sync_without_relocking()
            if not sync_result.success:
                raise RuntimeError(sync_result.message)
            _journal(operation, "CACHE_SYNCED", "CACHE_SYNCED")
            message = "Excel updated and website synchronised successfully."
            record_field_audit(_audit_rows(operation, serial, safe_changes, original_display, current, "Success", message))
            _journal(operation, "COMMITTED", "COMMITTED")
            _finish_journal(operation, "Success", message)
            return OperationResult(
                "updated", message, operation_id=operation["operation_id"], serial_number=serial,
                serial_numbers=(serial,), changed_fields=tuple(safe_changes), backup_file=str(backup or ""),
                cache_sync_status=sync_result.status, warnings=tuple(sync_result.warnings),
                already_applied=tuple(conflict_check["already_applied"]),
            )
    except Exception as error:
        message = str(error)
        try:
            _journal(operation, "RECOVERY_REQUIRED" if source_replaced else "FAILED", "RECOVERY_REQUIRED" if source_replaced else "FAILED", error=message)
            _finish_journal(operation, "Partial" if source_replaced else "Failed", message, clear_active=not source_replaced)
        except OSError:
            pass
        return OperationResult(
            "partial" if source_replaced else "failed",
            ("Excel was updated, but website refresh failed. Recovery can safely refresh the cache. " + message) if source_replaced else message,
            operation_id=operation["operation_id"], serial_number=serial, backup_file=str(backup or ""),
        )
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def execute_batch_updates(
    *, updates: Sequence[Mapping[str, Any]], user: str, session_id: str, source_page: str,
) -> OperationResult:
    """Apply all unit changes in one workbook replacement; any conflict stops the batch."""
    cleaned = [item for item in updates if clean_text(item.get("serial_number", ""))]
    if not cleaned:
        return OperationResult("no_changes", "No batch changes were supplied.")
    serials = [clean_text(item["serial_number"]) for item in cleaned]
    operation = _new_operation("BATCH_UPDATE", user, session_id, serials, source_page)
    temp_path: Path | None = None
    backup: Path | None = None
    source_replaced = False
    try:
        with operation_lock({
            "operation_id": operation["operation_id"], "operation_type": "BATCH_UPDATE",
            "user": user, "session_id": session_id, "serial_numbers": serials,
        }, EXCEL_OPERATION_LOCK_FILE):
            _journal(operation, "LOCK_ACQUIRED", "LOCK_ACQUIRED")
            temp_path = _make_temporary_workbook_path(Path(EXCEL_FILE))
            operation["temp_file"] = str(temp_path)
            _copy_stable_workbook(Path(EXCEL_FILE), temp_path)
            wb = load_workbook(temp_path, data_only=False)
            baseline = _capture_baseline(wb, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW)
            ws = wb[EXCEL_SHEET]
            headers = _header_map(ws, EXCEL_HEADER_ROW)
            serial_col = _required_column(headers, "SERIAL NUMBER")
            staged: list[tuple[str, int, dict[str, Any], dict[str, str], dict[str, str], dict[str, str]]] = []
            all_conflicts: dict[str, Any] = {}
            all_already: list[str] = []

            for item in cleaned:
                serial = clean_text(item["serial_number"])
                desired = _normalise_changes(item.get("desired_values", {}))
                original_raw = item.get("original_values", {})
                row = _find_serial_row(ws, serial_column=serial_col, data_start_row=EXCEL_DATA_START_ROW, serial_number=serial)
                current: dict[str, str] = {}
                original: dict[str, str] = {}
                desired_display: dict[str, str] = {}
                for field, value in desired.items():
                    col = _required_column(headers, APP_TO_EXCEL_COLUMNS[field])
                    if field == "Remarks":
                        current[field] = _read_combined_remarks(
                            ws,
                            row_number=row,
                            serial_column=serial_col,
                            remarks_column=col,
                        )
                    else:
                        current[field] = _read_app_value(
                            ws.cell(row=row, column=col), field
                        )
                    original[field] = _display_value(field, original_raw.get(field, ""))
                    desired_display[field] = _display_value(field, value)
                check = detect_field_conflicts(original, current, desired_display)
                if check["conflicts"]:
                    all_conflicts[serial] = check["conflicts"]
                all_already.extend(f"{serial}:{field}" for field in check["already_applied"])
                safe = {field: desired[field] for field in check["safe_changes"]}
                staged.append((serial, row, safe, original, current, desired_display))

            _journal(operation, "CONFLICT_CHECKED", "CONFLICT_CHECKED")
            if all_conflicts:
                message = "The batch was not saved because one or more fields have conflicts."
                conflict_rows: list[dict[str, Any]] = []
                for serial, fields in all_conflicts.items():
                    for field, values in fields.items():
                        conflict_rows.append({
                            "Operation ID": operation["operation_id"], "User": user, "Session ID": session_id,
                            "Operation Type": "BATCH_UPDATE", "Source Page": source_page, "Serial Number": serial,
                            "Field": field, "Original Value": values["original"], "Current Value": values["current"],
                            "Desired Value": values["desired"], "Result": "Conflict", "Message": message,
                        })
                record_conflicts(conflict_rows)
                record_field_audit(conflict_rows)
                wb.close(); temp_path.unlink(missing_ok=True); temp_path = None
                _journal(operation, "CONFLICT_DETECTED", "CONFLICT_DETECTED")
                _finish_journal(operation, "Conflict", message)
                return OperationResult("conflict", message, operation_id=operation["operation_id"], serial_numbers=tuple(serials), conflicts=all_conflicts)

            actual_count = sum(len(item[2]) for item in staged)
            if actual_count == 0:
                message = "The requested values were already present in Excel. No additional update was required."
                wb.close(); temp_path.unlink(missing_ok=True); temp_path = None
                _finish_journal(operation, "Already Applied", message)
                return OperationResult("already_applied", message, operation_id=operation["operation_id"], serial_numbers=tuple(serials), already_applied=tuple(all_already))

            staging = wb.create_sheet(STAGING_SHEET_NAME); staging.sheet_state = "hidden"
            staging.append(["Operation ID", "Serial Number", "Field", "Excel Row", "Original", "Current", "Desired"])
            expected: dict[str, dict[str, Any]] = {}
            allowed_continuation_rows: set[int] = set()
            for serial, row, safe, original, current, desired_display in staged:
                expected[serial] = safe
                for field, value in safe.items():
                    staging.append([operation["operation_id"], serial, field, row, original.get(field, ""), current.get(field, ""), desired_display.get(field, "")])
                    col = _required_column(headers, APP_TO_EXCEL_COLUMNS[field])
                    if field == "Remarks":
                        allowed_continuation_rows.update(
                            _write_combined_remarks(
                                ws,
                                row_number=row,
                                serial_column=serial_col,
                                remarks_column=col,
                                value=value,
                            )
                        )
                    else:
                        _write_cell(ws.cell(row=row, column=col), field, value)
            wb.save(temp_path); wb.close()
            _journal(operation, "CHANGES_APPLIED", "CHANGES_APPLIED")

            first = load_workbook(temp_path, data_only=False)
            _validate_structure(first, baseline=baseline, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW, staging_expected=True, allowed_continuation_rows=allowed_continuation_rows)
            for serial, changes in expected.items():
                if changes:
                    _validate_target_values(first, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW, serial_number=serial, expected_changes=changes)
            del first[STAGING_SHEET_NAME]; first.save(temp_path); first.close()
            second = load_workbook(temp_path, data_only=False)
            _validate_structure(second, baseline=baseline, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW, staging_expected=False, allowed_continuation_rows=allowed_continuation_rows)
            for serial, changes in expected.items():
                if changes:
                    _validate_target_values(second, sheet_name=EXCEL_SHEET, header_row=EXCEL_HEADER_ROW, data_start_row=EXCEL_DATA_START_ROW, serial_number=serial, expected_changes=changes)
            second.close(); _journal(operation, "TEMP_VALIDATED", "TEMP_VALIDATED")

            backup = _create_excel_backup(Path(EXCEL_FILE), Path(EXCEL_BACKUP_DIR), MAX_EXCEL_BACKUPS)
            operation["backup_file"] = str(backup); _journal(operation, "BACKUP_CREATED", "BACKUP_CREATED")
            os.replace(temp_path, EXCEL_FILE); temp_path = None; source_replaced = True
            _journal(operation, "SOURCE_REPLACED", "SOURCE_REPLACED")
            sync_result = _sync_without_relocking()
            if not sync_result.success:
                raise RuntimeError(sync_result.message)
            _journal(operation, "CACHE_SYNCED", "CACHE_SYNCED")
            message = f"Updated {len(serials)} AED unit(s) in one Excel transaction."
            audit_rows: list[dict[str, Any]] = []
            for serial, _, safe, original, current, _ in staged:
                audit_rows.extend(_audit_rows(operation, serial, safe, original, current, "Success", message))
            record_field_audit(audit_rows)
            _journal(operation, "COMMITTED", "COMMITTED"); _finish_journal(operation, "Success", message)
            changed_fields = tuple(sorted({field for _, _, safe, _, _, _ in staged for field in safe}))
            return OperationResult("updated", message, operation_id=operation["operation_id"], serial_numbers=tuple(serials), changed_fields=changed_fields, backup_file=str(backup), cache_sync_status=sync_result.status, warnings=tuple(sync_result.warnings))
    except Exception as error:
        message = str(error)
        try:
            _journal(operation, "RECOVERY_REQUIRED" if source_replaced else "FAILED", "RECOVERY_REQUIRED" if source_replaced else "FAILED", error=message)
            _finish_journal(operation, "Partial" if source_replaced else "Failed", message, clear_active=not source_replaced)
        except OSError:
            pass
        return OperationResult("partial" if source_replaced else "failed", message, operation_id=operation["operation_id"], serial_numbers=tuple(serials), backup_file=str(backup or ""))
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def _coerce_add_value(field: str, value: Any) -> Any:
    text = clean_text(value)
    if not text:
        return None
    if field == "Postal Code":
        if not text.isdigit() or len(text) > 6:
            raise ValueError("Postal Code must contain at most six digits.")
        return text.zfill(6)
    if field == "Lift Lobby":
        return _normalise_lift_lobby(text)
    if field in DATE_ADD_FIELDS:
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"{field} must be a real date.")
    return text


def execute_add_unit(*, values: Mapping[str, Any], user: str, session_id: str, source_page: str) -> OperationResult:
    serial = clean_text(values.get("Serial Number", ""))
    required = ["Serial Number", "Block / Locations", "Street Name", "Postal Code", "Model", "Next PM Date"]
    missing = [field for field in required if not clean_text(values.get(field, ""))]
    if missing:
        return OperationResult("failed", "Missing required fields: " + ", ".join(missing), serial_number=serial)
    operation = _new_operation("ADD_UNIT", user, session_id, [serial], source_page)
    temp_path: Path | None = None; backup: Path | None = None; source_replaced = False
    try:
        with operation_lock({"operation_id": operation["operation_id"], "operation_type": "ADD_UNIT", "user": user, "session_id": session_id, "serial_numbers": [serial]}, EXCEL_OPERATION_LOCK_FILE):
            _journal(operation, "LOCK_ACQUIRED", "LOCK_ACQUIRED")
            temp_path = _make_temporary_workbook_path(Path(EXCEL_FILE)); operation["temp_file"] = str(temp_path)
            _copy_stable_workbook(Path(EXCEL_FILE), temp_path)
            wb = load_workbook(temp_path, data_only=False); ws = wb[EXCEL_SHEET]
            headers = _header_map(ws, EXCEL_HEADER_ROW); serial_col = _required_column(headers, "SERIAL NUMBER")
            existing = [value.casefold() for _, value in _serial_rows(ws, serial_column=serial_col, data_start_row=EXCEL_DATA_START_ROW)]
            if serial.casefold() in existing:
                raise ValueError(f"Serial Number already exists: {serial}")
            valid_rows = [row for row, _ in _serial_rows(ws, serial_column=serial_col, data_start_row=EXCEL_DATA_START_ROW)]
            if not valid_rows:
                raise ValueError("No existing AED row is available for style copying.")
            source_row = valid_rows[-1]; target_row = source_row + 1
            # Do not overwrite a continuation or meaningful row.
            while any(clean_text(ws.cell(target_row, col).value) for col in range(1, ws.max_column + 1)):
                target_row += 1
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
            for col in range(1, ws.max_column + 1):
                src = ws.cell(source_row, col); dst = ws.cell(target_row, col)
                if src.has_style:
                    dst._style = copy(src._style)
                if src.number_format:
                    dst.number_format = src.number_format
                dst.font = copy(src.font); dst.fill = copy(src.fill); dst.border = copy(src.border); dst.alignment = copy(src.alignment); dst.protection = copy(src.protection)

            staging = wb.create_sheet(STAGING_SHEET_NAME); staging.sheet_state = "hidden"
            staging.append(["Operation ID", "Operation", "Serial Number", "Target Row"])
            staging.append([operation["operation_id"], "ADD_UNIT", serial, target_row])
            for field, header in ADD_TO_EXCEL_COLUMNS.items():
                if field not in values:
                    continue
                col = _required_column(headers, header)
                value = _coerce_add_value(field, values.get(field))
                cell = ws.cell(target_row, col); cell.value = value
                if field == "Postal Code": cell.number_format = "@"
                if field in DATE_ADD_FIELDS: cell.number_format = "dd/mm/yyyy"
            wb.save(temp_path); wb.close()
            check = load_workbook(temp_path, data_only=False); cws = check[EXCEL_SHEET]
            if clean_text(cws.cell(target_row, serial_col).value) != serial:
                raise ValueError("Temporary workbook validation failed for the new Serial Number.")
            del check[STAGING_SHEET_NAME]; check.save(temp_path); check.close()
            final = load_workbook(temp_path, data_only=False)
            if STAGING_SHEET_NAME in final.sheetnames or clean_text(final[EXCEL_SHEET].cell(target_row, serial_col).value) != serial:
                raise ValueError("Final temporary workbook validation failed.")
            final.close(); _journal(operation, "TEMP_VALIDATED", "TEMP_VALIDATED")
            backup = _create_excel_backup(Path(EXCEL_FILE), Path(EXCEL_BACKUP_DIR), MAX_EXCEL_BACKUPS)
            operation["backup_file"] = str(backup); _journal(operation, "BACKUP_CREATED", "BACKUP_CREATED")
            os.replace(temp_path, EXCEL_FILE); temp_path = None; source_replaced = True; _journal(operation, "SOURCE_REPLACED", "SOURCE_REPLACED")
            sync_result = _sync_without_relocking()
            if not sync_result.success: raise RuntimeError(sync_result.message)
            _journal(operation, "CACHE_SYNCED", "CACHE_SYNCED")
            message = f"Added {serial} to Excel and refreshed the website."
            record_field_audit([{
                "Operation ID": operation["operation_id"], "User": user, "Session ID": session_id,
                "Operation Type": "ADD_UNIT", "Source Page": source_page, "Serial Number": serial,
                "Field": "Record", "Original Value": "", "Current Value": "", "Desired Value": "New AED",
                "Result": "Success", "Message": message,
            }])
            _journal(operation, "COMMITTED", "COMMITTED"); _finish_journal(operation, "Success", message)
            return OperationResult("added", message, operation_id=operation["operation_id"], serial_number=serial, serial_numbers=(serial,), backup_file=str(backup), cache_sync_status=sync_result.status, warnings=tuple(sync_result.warnings))
    except Exception as error:
        message = str(error)
        try:
            _journal(operation, "RECOVERY_REQUIRED" if source_replaced else "FAILED", "RECOVERY_REQUIRED" if source_replaced else "FAILED", error=message); _finish_journal(operation, "Partial" if source_replaced else "Failed", message, clear_active=not source_replaced)
        except OSError: pass
        return OperationResult("partial" if source_replaced else "failed", message, operation_id=operation["operation_id"], serial_number=serial, backup_file=str(backup or ""))
    finally:
        if temp_path is not None: temp_path.unlink(missing_ok=True)


def _append_lifecycle(row: Mapping[str, Any]) -> None:
    path = Path(AED_LIFECYCLE_FILE); path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists() or path.stat().st_size == 0
    with path.open("a", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=LIFECYCLE_COLUMNS)
        if write_header: writer.writeheader()
        writer.writerow({column: row.get(column, "") for column in LIFECYCLE_COLUMNS})


def execute_deactivate_unit(*, serial_number: str, user: str, session_id: str, source_page: str, reason: str = "") -> OperationResult:
    serial = clean_text(serial_number)
    operation = _new_operation("DEACTIVATE_UNIT", user, session_id, [serial], source_page)
    if not serial or not clean_text(user):
        return OperationResult("failed", "Serial Number and audit user are required.", serial_number=serial)
    try:
        with operation_lock({"operation_id": operation["operation_id"], "operation_type": "DEACTIVATE_UNIT", "user": user, "session_id": session_id, "serial_numbers": [serial]}, EXCEL_OPERATION_LOCK_FILE):
            _journal(operation, "LOCK_ACQUIRED", "LOCK_ACQUIRED")
            # Verify the unit still exists in the current official workbook.
            wb = load_workbook(EXCEL_FILE, read_only=True, data_only=False); ws = wb[EXCEL_SHEET]
            headers = _header_map(ws, EXCEL_HEADER_ROW); serial_col = _required_column(headers, "SERIAL NUMBER")
            _find_serial_row(ws, serial_column=serial_col, data_start_row=EXCEL_DATA_START_ROW, serial_number=serial); wb.close()
            _append_lifecycle({
                "Timestamp": _now_iso(), "Operation ID": operation["operation_id"], "User": user,
                "Session ID": session_id, "Serial Number": serial, "Status": "Inactive",
                "Reason": clean_text(reason), "Source Page": source_page,
            })
            message = f"{serial} was marked inactive. Its Excel row and service history were preserved."
            record_field_audit([{
                "Operation ID": operation["operation_id"], "User": user, "Session ID": session_id,
                "Operation Type": "DEACTIVATE_UNIT", "Source Page": source_page, "Serial Number": serial,
                "Field": "Lifecycle Status", "Original Value": "Active", "Current Value": "Active",
                "Desired Value": "Inactive", "Result": "Success", "Message": message,
            }])
            _journal(operation, "COMMITTED", "COMMITTED"); _finish_journal(operation, "Success", message)
            return OperationResult("deactivated", message, operation_id=operation["operation_id"], serial_number=serial, serial_numbers=(serial,))
    except Exception as error:
        message = str(error)
        try: _journal(operation, "FAILED", "FAILED", error=message); _finish_journal(operation, "Failed", message)
        except OSError: pass
        return OperationResult("failed", message, operation_id=operation["operation_id"], serial_number=serial)


def load_latest_lifecycle_status() -> dict[str, str]:
    path = Path(AED_LIFECYCLE_FILE)
    if not path.exists() or path.stat().st_size == 0:
        return {}
    latest: dict[str, str] = {}
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            serial = clean_text(row.get("Serial Number", "")).casefold()
            if serial:
                latest[serial] = clean_text(row.get("Status", "")) or "Active"
    return latest
