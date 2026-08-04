from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from services.excel_write_service import (
    STAGING_SHEET_NAME,
    load_excel_write_history,
    update_existing_unit_in_excel,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REAL_WORKBOOK = PROJECT_ROOT / "external_data" / "IB_list_TEST.xlsx"
REAL_CACHE = PROJECT_ROOT / "aed_data.csv"


def _paths(tmp_path: Path) -> dict[str, Path]:
    paths = {
        "excel": tmp_path / "external" / "IB.xlsx",
        "cache": tmp_path / "aed_data.csv",
        "write_lock": tmp_path / "data" / "write.lock",
        "sync_lock": tmp_path / "data" / "sync.lock",
        "state": tmp_path / "data" / "sync.json",
        "history": tmp_path / "data" / "write_history.csv",
        "excel_backups": tmp_path / "backups" / "excel",
        "cache_backups": tmp_path / "backups" / "cache",
        "temp": tmp_path / "temp",
    }
    for key in [
        "excel",
        "write_lock",
        "sync_lock",
        "state",
        "history",
    ]:
        paths[key].parent.mkdir(parents=True, exist_ok=True)
    paths["excel_backups"].mkdir(parents=True, exist_ok=True)
    paths["cache_backups"].mkdir(parents=True, exist_ok=True)
    paths["temp"].mkdir(parents=True, exist_ok=True)
    return paths


def _call(paths: dict[str, Path], **overrides):
    kwargs = {
        "serial_number": "X18K075125",
        "changes": {"Block / Locations": "128A"},
        "original_values": {"Block / Locations": "128"},
        "user": "Zihan",
        "source_page": "AED Management",
        "excel_file": paths["excel"],
        "excel_sheet": "Sheet1",
        "write_lock_file": paths["write_lock"],
        "excel_backup_dir": paths["excel_backups"],
        "write_history_file": paths["history"],
        "cache_file": paths["cache"],
        "sync_state_file": paths["state"],
        "sync_lock_file": paths["sync_lock"],
        "cache_backup_dir": paths["cache_backups"],
        "temp_dir": paths["temp"],
        "preserve_cache_only_units": False,
    }
    kwargs.update(overrides)
    return update_existing_unit_in_excel(**kwargs)


def _copy_real_files(paths: dict[str, Path]) -> None:
    shutil.copy2(REAL_WORKBOOK, paths["excel"])
    shutil.copy2(REAL_CACHE, paths["cache"])


def _make_minimal_workbook(path: Path, serials: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    headers = [
        "SERIAL NUMBER",
        "Block / Locations",
        "Street Name",
        "Postal Code",
        "Next PM Due",
        "Remarks",
    ]
    worksheet.append(headers)
    worksheet.append([None] * len(headers))
    for index, serial in enumerate(serials, start=1):
        worksheet.append(
            [serial, str(100 + index), "Test Street", "123456", None, "Test"]
        )
    workbook.create_sheet("Reference")
    workbook.save(path)


def test_safe_writeback_updates_excel_cache_history_and_backup(tmp_path: Path):
    paths = _paths(tmp_path)
    _copy_real_files(paths)

    before = load_workbook(paths["excel"], data_only=False)
    original_continuations = (before["Sheet1"]["Y4"].value, before["Sheet1"]["Y6"].value)
    original_merged = sorted(str(item) for item in before["Sheet1"].merged_cells.ranges)
    original_row_heights = {
        index: dimension.height
        for index, dimension in before["Sheet1"].row_dimensions.items()
        if dimension.height is not None
    }
    original_style_ids = {
        cell.coordinate: cell.style_id
        for row in before["Sheet1"].iter_rows()
        for cell in row
    }
    before.close()

    result = _call(
        paths,
        changes={
            "Block / Locations": "128A",
            "Next PM Date": "2026-10-15",
        },
        original_values={
            "Block / Locations": "128",
            "Next PM Date": "01-07-2027",
        },
    )

    assert result.success
    assert set(result.changed_fields) == {"Block / Locations", "Next PM Date"}

    workbook = load_workbook(paths["excel"], data_only=False)
    worksheet = workbook["Sheet1"]
    assert worksheet["G7"].value == "128A"
    assert worksheet["U7"].value.date().isoformat() == "2026-10-15"
    assert worksheet["U7"].number_format == "dd/mm/yyyy"
    assert STAGING_SHEET_NAME not in workbook.sheetnames
    assert (worksheet["Y4"].value, worksheet["Y6"].value) == original_continuations
    assert sorted(str(item) for item in worksheet.merged_cells.ranges) == original_merged
    assert {
        index: dimension.height
        for index, dimension in worksheet.row_dimensions.items()
        if dimension.height is not None
    } == original_row_heights
    current_style_ids = {
        cell.coordinate: cell.style_id
        for row in worksheet.iter_rows()
        for cell in row
        if cell.coordinate != "U7"
    }
    expected_style_ids = {
        coordinate: style_id
        for coordinate, style_id in original_style_ids.items()
        if coordinate != "U7"
    }
    assert current_style_ids == expected_style_ids
    workbook.close()

    cache = pd.read_csv(paths["cache"], dtype=str, keep_default_na=False)
    row = cache.loc[cache["Serial Number"].eq("X18K075125")].iloc[0]
    assert row["Block / Locations"] == "128A"
    assert row["Location"] == "Blk 128A Bukit Merah View"
    assert row["Next PM Date"] == "15-10-2026"

    backups = list(paths["excel_backups"].glob("IB_*.xlsx"))
    assert len(backups) == 1
    backup = load_workbook(backups[0], data_only=False)
    assert backup["Sheet1"]["G7"].value == 128
    backup.close()

    history = load_excel_write_history(paths["history"])
    assert len(history) == 2
    assert {item["Field"] for item in history} == {
        "Block / Locations",
        "Next PM Date",
    }
    assert all(item["Result"] == "Success" for item in history)


def test_postal_change_clears_old_coordinates_in_cache(tmp_path: Path):
    paths = _paths(tmp_path)
    _copy_real_files(paths)

    result = _call(
        paths,
        changes={"Postal Code": "560123"},
        original_values={"Postal Code": "150128"},
    )

    assert result.success
    workbook = load_workbook(paths["excel"], data_only=False)
    cell = workbook["Sheet1"]["I7"]
    assert cell.value == "560123"
    assert cell.number_format == "@"
    workbook.close()

    cache = pd.read_csv(paths["cache"], dtype=str, keep_default_na=False)
    row = cache.loc[cache["Serial Number"].eq("X18K075125")].iloc[0]
    assert row["Postal Code"] == "560123"
    assert row["Latitude"] == ""
    assert row["Longitude"] == ""
    assert "pending" in row["Geocoding Status"].casefold()


def test_no_change_does_not_modify_excel_or_create_backup(tmp_path: Path):
    paths = _paths(tmp_path)
    _copy_real_files(paths)
    before = paths["excel"].stat().st_mtime_ns

    result = _call(
        paths,
        changes={"Block / Locations": "128"},
        original_values={"Block / Locations": "128"},
    )

    assert result.status == "no_changes"
    assert paths["excel"].stat().st_mtime_ns == before
    assert not list(paths["excel_backups"].glob("*.xlsx"))


def test_read_only_and_unsupported_fields_are_rejected(tmp_path: Path):
    paths = _paths(tmp_path)
    _copy_real_files(paths)

    read_only = _call(
        paths,
        changes={"Location": "Wrong"},
        original_values={"Location": "Old"},
    )
    unsupported = _call(
        paths,
        changes={"Unknown Field": "AED Plus"},
        original_values={"Unknown Field": ""},
    )
    supported_model = _call(
        paths,
        changes={"Model": "AED Plus Test"},
        original_values={"Model": "ZOLL AED Plus"},
    )

    assert read_only.status == "failed"
    assert "read-only" in read_only.message
    assert unsupported.status == "failed"
    assert "Unsupported" in unsupported.message
    assert supported_model.status == "updated"
    workbook = load_workbook(paths["excel"], data_only=False)
    assert workbook["Sheet1"]["B7"].value == "AED Plus Test"
    workbook.close()


def test_missing_and_duplicate_serial_do_not_replace_workbook(tmp_path: Path):
    paths = _paths(tmp_path)
    _make_minimal_workbook(paths["excel"], ["DUP001", "DUP001"])
    pd.DataFrame([{"Serial Number": "DUP001"}]).to_csv(
        paths["cache"], index=False, encoding="utf-8-sig"
    )
    before = paths["excel"].read_bytes()

    duplicate = _call(
        paths,
        serial_number="DUP001",
        changes={"Block / Locations": "999"},
        original_values={"Block / Locations": "101"},
    )
    missing = _call(
        paths,
        serial_number="NOT-EXIST-001",
        changes={"Block / Locations": "999"},
        original_values={"Block / Locations": ""},
    )

    assert duplicate.status == "failed"
    assert "Duplicate Serial Number" in duplicate.message
    assert missing.status == "failed"
    assert "Unit not found" in missing.message
    assert paths["excel"].read_bytes() == before


def test_existing_write_lock_blocks_update_and_is_preserved(tmp_path: Path):
    paths = _paths(tmp_path)
    _copy_real_files(paths)
    paths["write_lock"].write_text(
        '{"user":"Other User","started_at":"2026-08-01T12:00:00+08:00"}',
        encoding="utf-8",
    )
    before = paths["excel"].read_bytes()

    result = _call(paths)

    assert result.status == "failed"
    assert "already in progress" in result.message
    assert paths["write_lock"].exists()
    assert paths["excel"].read_bytes() == before



def test_successful_write_preserves_other_worksheets(tmp_path: Path):
    paths = _paths(tmp_path)
    _make_minimal_workbook(paths["excel"], ["ONLY001"])
    pd.DataFrame(
        [
            {
                "Serial Number": "ONLY001",
                "Block / Locations": "101",
                "Street Name": "Test Street",
                "Location": "Blk 101 Test Street",
                "Postal Code": "123456",
                "Next PM Date": "",
            }
        ]
    ).to_csv(paths["cache"], index=False, encoding="utf-8-sig")

    before = load_workbook(paths["excel"], data_only=False)
    before["Reference"]["A1"] = "KEEP ME"
    before.save(paths["excel"])
    before.close()

    result = _call(
        paths,
        serial_number="ONLY001",
        changes={"Street Name": "Updated Street"},
        original_values={"Street Name": "Test Street"},
    )

    assert result.success
    workbook = load_workbook(paths["excel"], data_only=False)
    assert workbook.sheetnames == ["Sheet1", "Reference"]
    assert workbook["Reference"]["A1"].value == "KEEP ME"
    assert workbook["Sheet1"]["C3"].value == "Updated Street"
    assert STAGING_SHEET_NAME not in workbook.sheetnames
    workbook.close()

def test_changed_by_is_required(tmp_path: Path):
    paths = _paths(tmp_path)
    _copy_real_files(paths)

    result = _call(paths, user="")

    assert result.status == "failed"
    assert "Changed By is required" in result.message
