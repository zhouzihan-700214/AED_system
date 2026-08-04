from __future__ import annotations

import json
import shutil
import socket
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from services.conflict_service import detect_field_conflicts
from services.excel_lock_service import ExcelOperationLock, inspect_lock, remove_confirmed_stale_lock
from services import audit_service
from services import excel_transaction_service as tx
from services import recovery_service
from services.excel_sync_service import SyncResult


PROJECT_EXCEL = Path(__file__).resolve().parents[1] / "external_data" / "IB_list_TEST.xlsx"
PROJECT_CACHE = Path(__file__).resolve().parents[1] / "aed_data.csv"


def _configure_transaction_paths(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> dict[str, Path]:
    paths = {
        "excel": tmp_path / "external_data" / "IB_list_TEST.xlsx",
        "cache": tmp_path / "aed_data.csv",
        "lock": tmp_path / "data" / "excel_operation.lock",
        "active": tmp_path / "data" / "active_transaction.json",
        "sync_state": tmp_path / "data" / "excel_sync_state.json",
        "excel_backup": tmp_path / "backups" / "excel",
        "cache_backup": tmp_path / "backups" / "aed_cache",
        "temp": tmp_path / "temp",
        "lifecycle": tmp_path / "data" / "aed_lifecycle_history.csv",
        "audit": tmp_path / "data" / "audit_history.csv",
        "conflict": tmp_path / "data" / "conflict_history.csv",
        "transactions": tmp_path / "data" / "transaction_history.csv",
    }
    for key in ("excel_backup", "cache_backup", "temp"):
        paths[key].mkdir(parents=True, exist_ok=True)
    paths["excel"].parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(PROJECT_EXCEL, paths["excel"])
    shutil.copy2(PROJECT_CACHE, paths["cache"])
    mapping = {
        "EXCEL_FILE": paths["excel"], "AED_CACHE_FILE": paths["cache"],
        "EXCEL_OPERATION_LOCK_FILE": paths["lock"], "ACTIVE_TRANSACTION_FILE": paths["active"],
        "SYNC_STATE_FILE": paths["sync_state"], "EXCEL_BACKUP_DIR": paths["excel_backup"],
        "CACHE_BACKUP_DIR": paths["cache_backup"], "TEMP_DIR": paths["temp"],
        "AED_LIFECYCLE_FILE": paths["lifecycle"],
    }
    for name, value in mapping.items():
        monkeypatch.setattr(tx, name, value)
    monkeypatch.setattr(audit_service, "AUDIT_HISTORY_FILE", paths["audit"])
    monkeypatch.setattr(audit_service, "CONFLICT_HISTORY_FILE", paths["conflict"])
    monkeypatch.setattr(audit_service, "TRANSACTION_HISTORY_FILE", paths["transactions"])
    return paths


def _excel_value(path: Path, serial: str, header: str):
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Sheet1"]
    headers = {str(sheet.cell(1, column).value).strip(): column for column in range(1, sheet.max_column + 1)}
    serial_col = headers["SERIAL NUMBER"]
    for row in range(3, sheet.max_row + 1):
        if str(sheet.cell(row, serial_col).value or "").strip() == serial:
            value = sheet.cell(row, headers[header]).value
            workbook.close()
            return value
    workbook.close()
    raise AssertionError(f"Serial not found: {serial}")


def test_conflict_service_three_outcomes() -> None:
    result = detect_field_conflicts(
        {"A": "old", "B": "old", "C": "old"},
        {"A": "old", "B": "new", "C": "other"},
        {"A": "new", "B": "new", "C": "wanted"},
    )
    assert result["safe_changes"] == {"A": "new"}
    assert result["already_applied"] == {"B": "new"}
    assert result["conflicts"]["C"]["current"] == "other"


def test_atomic_lock_blocks_second_operation(tmp_path: Path) -> None:
    path = tmp_path / "excel_operation.lock"
    first = ExcelOperationLock(path)
    first.acquire({"operation_id": "one", "operation_type": "TEST", "user": "A"})
    with pytest.raises(RuntimeError):
        ExcelOperationLock(path).acquire({"operation_id": "two", "operation_type": "TEST", "user": "B"})
    first.release()
    assert not path.exists()


def test_confirmed_stale_same_host_lock_is_removable(tmp_path: Path) -> None:
    path = tmp_path / "excel_operation.lock"
    path.write_text(json.dumps({
        "operation_id": "old", "hostname": socket.gethostname(), "process_id": 99999999,
        "started_at": "2026-01-01T00:00:00+08:00",
    }), encoding="utf-8")
    assert inspect_lock(path)["confirmed_stale"] is True
    assert remove_confirmed_stale_lock(path) is True
    assert not path.exists()


def test_unit_update_and_same_field_conflict(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    paths = _configure_transaction_paths(monkeypatch, tmp_path)
    result = tx.execute_unit_update(
        serial_number="X18K075125",
        desired_values={"Block / Locations": "128A"},
        original_values={"Block / Locations": "128"},
        user="Zihan", session_id="one", source_page="Test",
    )
    assert result.success
    assert _excel_value(paths["excel"], "X18K075125", "Block / Locations") == "128A"
    stale = tx.execute_unit_update(
        serial_number="X18K075125",
        desired_values={"Block / Locations": "128B"},
        original_values={"Block / Locations": "128"},
        user="Zihan", session_id="two", source_page="Test",
    )
    assert stale.status == "conflict"
    assert _excel_value(paths["excel"], "X18K075125", "Block / Locations") == "128A"


def test_different_field_after_other_update_can_merge(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    paths = _configure_transaction_paths(monkeypatch, tmp_path)
    first = tx.execute_unit_update(
        serial_number="X18K075125", desired_values={"Block / Locations": "128A"},
        original_values={"Block / Locations": "128"}, user="Zihan", session_id="one", source_page="Test",
    )
    assert first.success
    second = tx.execute_unit_update(
        serial_number="X18K075125", desired_values={"Next PM Date": "15-10-2027"},
        original_values={"Next PM Date": "01-07-2027"}, user="Zihan", session_id="two", source_page="Test",
    )
    assert second.success
    assert _excel_value(paths["excel"], "X18K075125", "Block / Locations") == "128A"
    assert str(_excel_value(paths["excel"], "X18K075125", "Next PM Due")).startswith("2027-10-15")


def test_batch_conflict_is_all_or_nothing(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    paths = _configure_transaction_paths(monkeypatch, tmp_path)
    # External/current change makes the first update stale.
    workbook = load_workbook(paths["excel"]); sheet = workbook["Sheet1"]
    sheet["G7"] = "128A"; workbook.save(paths["excel"]); workbook.close()
    result = tx.execute_batch_updates(
        updates=[
            {"serial_number": "X18K075125", "original_values": {"Block / Locations": "128"}, "desired_values": {"Block / Locations": "128B"}},
            {"serial_number": "X18K075198", "original_values": {"Next PM Date": "01-07-2027"}, "desired_values": {"Next PM Date": "02-11-2027"}},
        ],
        user="Zihan", session_id="batch", source_page="PM Planning",
    )
    assert result.status == "conflict"
    assert str(_excel_value(paths["excel"], "X18K075198", "Next PM Due")).startswith("2027-07-01")


def test_add_unit_appends_and_preserves_staging_cleanup(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    paths = _configure_transaction_paths(monkeypatch, tmp_path)
    result = tx.execute_add_unit(
        values={
            "Serial Number": "TEST-STAGE4-001", "Model": "AED Plus",
            "Block / Locations": "100", "Street Name": "Test Street",
            "Postal Code": "012345", "Next PM Date": "15-12-2027",
        },
        user="Zihan", session_id="add", source_page="AED Management",
    )
    assert result.status == "added"
    assert _excel_value(paths["excel"], "TEST-STAGE4-001", "Postal Code") == "012345"
    workbook = load_workbook(paths["excel"])
    assert "__STAGING_UPDATE__" not in workbook.sheetnames
    workbook.close()


def test_deactivation_preserves_excel_row(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    paths = _configure_transaction_paths(monkeypatch, tmp_path)
    result = tx.execute_deactivate_unit(
        serial_number="X18K075125", user="Zihan", session_id="deactivate",
        source_page="AED Management", reason="Transferred",
    )
    assert result.status == "deactivated"
    assert _excel_value(paths["excel"], "X18K075125", "Block / Locations") == 128
    status = tx.load_latest_lifecycle_status()
    assert status["x18k075125"] == "Inactive"


def test_recovery_after_source_replaced_only_refreshes_cache(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    active = tmp_path / "active_transaction.json"
    lock = tmp_path / "excel_operation.lock"
    active.write_text(json.dumps({
        "operation_id": "op-1", "operation_type": "UPDATE_UNIT", "status": "RECOVERY_REQUIRED",
        "user": "Zihan", "session_id": "s", "serial_numbers": ["AED-1"],
        "source_page": "Test", "started_at": "2026-08-01T12:00:00+08:00",
        "last_completed_step": "SOURCE_REPLACED", "temp_file": None,
    }), encoding="utf-8")
    monkeypatch.setattr(recovery_service, "ACTIVE_TRANSACTION_FILE", active)
    monkeypatch.setattr(recovery_service, "EXCEL_OPERATION_LOCK_FILE", lock)
    monkeypatch.setattr(recovery_service, "_sync_without_relocking", lambda: SyncResult(
        status="synced", message="ok", source_exists=True, changed=True, row_count=1
    ))
    recorded = []
    monkeypatch.setattr(recovery_service, "record_transaction", lambda **row: recorded.append(row))
    result = recovery_service.recover_incomplete_transaction()
    assert result["status"] == "recovered"
    assert not active.exists()
    assert recorded[0]["Result"] == "Recovered"


def test_full_detail_update_writes_extended_fields_and_consolidates_remarks(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    paths = _configure_transaction_paths(monkeypatch, tmp_path)
    cache = pd.read_csv(paths["cache"], dtype=str, keep_default_na=False)
    original_remarks = cache.loc[
        cache["Serial Number"].eq("X16B816938"), "Remarks"
    ].iloc[0]

    result = tx.execute_unit_update(
        serial_number="X16B816938",
        desired_values={
            "Model": "ZOLL AED Plus Updated",
            "Level": "2",
            "Lift Lobby": "Lift Lobby A",
            "Remarks": "Consolidated website remarks",
            "Repaired?": "Yes",
        },
        original_values={
            "Model": "X17H943760",
            "Level": "",
            "Lift Lobby": "Lift Lobby B",
            "Remarks": original_remarks,
            "Repaired?": "",
        },
        user="Zihan",
        session_id="full-detail",
        source_page="AED Management Full Details",
    )

    assert result.success
    workbook = load_workbook(paths["excel"], data_only=False)
    sheet = workbook["Sheet1"]
    assert sheet["B3"].value == "ZOLL AED Plus Updated"
    assert str(sheet["K3"].value) == "2"
    assert sheet["J3"].value == "A"
    assert sheet["Y3"].value == "Consolidated website remarks"
    assert sheet["Y4"].value is None
    assert sheet["AC3"].value == "Yes"
    assert "__STAGING_UPDATE__" not in workbook.sheetnames
    workbook.close()

    refreshed = pd.read_csv(paths["cache"], dtype=str, keep_default_na=False)
    row = refreshed.loc[refreshed["Serial Number"].eq("X16B816938")].iloc[0]
    assert row["Remarks"] == "Consolidated website remarks"
    assert row["Level"] == "2"
    assert row["Lift Lobby"] == "Lift Lobby A"


def test_add_unit_accepts_complete_business_record(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    paths = _configure_transaction_paths(monkeypatch, tmp_path)
    result = tx.execute_add_unit(
        values={
            "Serial Number": "TEST-FULL-001",
            "Installation Date": "01-08-2026",
            "Model": "ZOLL AED Plus",
            "Installed Phase / Month": "Phase 3 / Aug 2026",
            "PO Number": "PO-1001",
            "Zone": "Central",
            "Block / Locations": "100",
            "Street Name": "Test Street",
            "Postal Code": "012345",
            "Level": "1",
            "Lift Lobby": "Lift Lobby C",
            "Adult Pads Replacement Date": "01-08-2026",
            "Adult Pads Expiry Date": "01-08-2028",
            "Adult Pads Lot Number": "ADULT-1",
            "Pediatric Pads Replacement Date": "01-08-2026",
            "Pediatric Pads Expiry Date": "01-08-2028",
            "Pediatric Pads Lot Number": "PED-1",
            "Battery Replacement History": "01-08-2026",
            "Battery Expiry Date": "01-08-2030",
            "PM Completed Date": "01-08-2026",
            "Next PM Date": "01-08-2027",
            "Job Type": "Commissioning",
            "Last Done By": "Zihan",
            "Service Report e-SR": "e-SR-TEST",
            "Remarks": "Complete add form test",
            "Repaired?": "Not applicable",
        },
        user="Zihan",
        session_id="add-full",
        source_page="AED Management Add",
    )

    assert result.status == "added"
    assert _excel_value(paths["excel"], "TEST-FULL-001", "RELATED OBJECTS") == "ZOLL AED Plus"
    assert _excel_value(paths["excel"], "TEST-FULL-001", "PO#") == "PO-1001"
    assert _excel_value(paths["excel"], "TEST-FULL-001", "Lift Lobby") == "C"
    assert _excel_value(paths["excel"], "TEST-FULL-001", "Adult CPR-D Padz Lot Number") == "ADULT-1"
    assert _excel_value(paths["excel"], "TEST-FULL-001", "Repaired?") == "Not applicable"
    assert _excel_value(paths["excel"], "TEST-FULL-001", "Remarks") == "Complete add form test"


def test_blank_website_date_matches_na_excel_value(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    paths = _configure_transaction_paths(monkeypatch, tmp_path)
    result = tx.execute_unit_update(
        serial_number="X16A810545",
        desired_values={"Adult Pads Replacement Date": "01-08-2026"},
        original_values={"Adult Pads Replacement Date": ""},
        user="Zihan",
        session_id="na-date",
        source_page="AED Management Full Details",
    )
    assert result.success
    value = _excel_value(
        paths["excel"], "X16A810545", "Adult CPR-D Padz Replacement Date"
    )
    assert str(value).startswith("2026-08-01")


def test_table_style_batch_updates_multiple_units_and_fields(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    paths = _configure_transaction_paths(monkeypatch, tmp_path)
    original_lot = str(
        _excel_value(paths["excel"], "X18K075125", "Adult CPR-D Padz Lot Number")
        or ""
    )
    result = tx.execute_batch_updates(
        updates=[
            {
                "serial_number": "X18K075125",
                "original_values": {
                    "Block / Locations": "128",
                    "Adult Pads Lot Number": original_lot,
                },
                "desired_values": {
                    "Block / Locations": "128A",
                    "Adult Pads Lot Number": "TABLE-LOT-1",
                },
            },
            {
                "serial_number": "X18K075198",
                "original_values": {"Repaired?": ""},
                "desired_values": {"Repaired?": "No"},
            },
        ],
        user="Zihan",
        session_id="table-batch",
        source_page="AED Management Table",
    )
    assert result.success
    assert _excel_value(paths["excel"], "X18K075125", "Block / Locations") == "128A"
    assert _excel_value(paths["excel"], "X18K075125", "Adult CPR-D Padz Lot Number") == "TABLE-LOT-1"
    assert _excel_value(paths["excel"], "X18K075198", "Repaired?") == "No"
