from __future__ import annotations

import io
import shutil
import sys
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.modules.setdefault("streamlit", MagicMock())

from services import (
    aed_repository,
    audit_service,
    excel_transaction_service as tx,
    issue_service,
    pm_service,
    service_record_service,
    system_state_service,
    unit_profile_service,
)
from services.excel_sync_service import SyncResult, sync_excel_to_cache
from services.pm_service import PM_PLAN_COLUMNS
from services.aed_field_schema import DETAIL_EDITABLE_COLUMNS
from views import pm_checklist

ROOT = Path(__file__).resolve().parents[1]
REAL_WORKBOOK = ROOT / "external_data" / "IB_list_TEST.xlsx"
REAL_CACHE = ROOT / "aed_data.csv"


def _paths(tmp_path: Path) -> dict[str, Path]:
    p = {
        "excel": tmp_path / "external_data" / "IB_list_TEST.xlsx",
        "cache": tmp_path / "aed_data.csv",
        "lock": tmp_path / "data" / "excel_operation.lock",
        "active": tmp_path / "data" / "active_transaction.json",
        "sync_state": tmp_path / "data" / "excel_sync_state.json",
        "excel_backup": tmp_path / "backups" / "excel",
        "cache_backup": tmp_path / "backups" / "aed_cache",
        "temp": tmp_path / "temp",
        "lifecycle": tmp_path / "data" / "aed_lifecycle_history.csv",
        "audit": tmp_path / "data" / "audit_history.csv",
        "conflict": tmp_path / "data" / "conflict_history.csv",
        "transactions": tmp_path / "data" / "transaction_history.csv",
        "pm": tmp_path / "pm_responses.csv",
        "plan": tmp_path / "pm_plan_records.csv",
        "issues": tmp_path / "issue_records.csv",
        "issue_history": tmp_path / "issue_history.csv",
        "attachments": tmp_path / "issue_attachments.csv",
        "resolutions": tmp_path / "issue_resolution_submissions.csv",
        "photos": tmp_path / "issue_photos",
        "map_state": tmp_path / "map_unit_state.csv",
        "map_status": tmp_path / "map_status_definitions.csv",
        "manual": tmp_path / "manual_service_records.csv",
    }
    for key in ("excel_backup", "cache_backup", "temp", "photos"):
        p[key].mkdir(parents=True, exist_ok=True)
    p["excel"].parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(REAL_WORKBOOK, p["excel"])
    shutil.copy2(REAL_CACHE, p["cache"])
    return p


def _configure(monkeypatch: pytest.MonkeyPatch, p: dict[str, Path]) -> None:
    mapping = {
        "EXCEL_FILE": p["excel"], "AED_CACHE_FILE": p["cache"],
        "EXCEL_OPERATION_LOCK_FILE": p["lock"], "ACTIVE_TRANSACTION_FILE": p["active"],
        "SYNC_STATE_FILE": p["sync_state"], "EXCEL_BACKUP_DIR": p["excel_backup"],
        "CACHE_BACKUP_DIR": p["cache_backup"], "TEMP_DIR": p["temp"],
        "AED_LIFECYCLE_FILE": p["lifecycle"],
    }
    for name, value in mapping.items():
        monkeypatch.setattr(tx, name, value)
    monkeypatch.setattr(audit_service, "AUDIT_HISTORY_FILE", p["audit"])
    monkeypatch.setattr(audit_service, "CONFLICT_HISTORY_FILE", p["conflict"])
    monkeypatch.setattr(audit_service, "TRANSACTION_HISTORY_FILE", p["transactions"])

    monkeypatch.setattr(pm_service, "PM_RESPONSES_FILE", p["pm"])
    monkeypatch.setattr(pm_service, "PM_PLAN_FILE", p["plan"])
    monkeypatch.setattr(pm_service, "AED_DATA_FILE", p["cache"])
    monkeypatch.setattr(pm_service, "AED_HISTORY_FILE", tmp_path_placeholder := p["audit"].with_name("aed_management_history.csv"))

    monkeypatch.setattr(pm_checklist, "PM_PLAN_FILE", p["plan"])
    monkeypatch.setattr(pm_checklist, "ISSUE_RECORD_FILE", p["issues"])
    monkeypatch.setattr(pm_checklist, "PM_RESPONSES_FILE", p["pm"], raising=False)

    # Imported default state paths in unit_color_service must be patched as well.
    from services import unit_color_service
    monkeypatch.setattr(unit_color_service, "MAP_UNIT_STATE_FILE", p["map_state"])
    monkeypatch.setattr(unit_color_service, "MAP_STATUS_FILE", p["map_status"])
    monkeypatch.setattr(pm_checklist, "sync_unit_from_issue_records", lambda issue_file, serial, clear_role="Completed": unit_color_service.sync_unit_from_issue_records(
        issue_file, serial, clear_role=clear_role, state_file=p["map_state"], status_file=p["map_status"]
    ))

    monkeypatch.setattr(aed_repository, "is_cloud_onedrive_enabled", lambda: False)


def _excel_row(path: Path, serial: str) -> tuple[object, int, dict[str, int]]:
    wb = load_workbook(path, data_only=False)
    ws = wb["Sheet1"]
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    serial_col = headers["SERIAL NUMBER"]
    row = next(r for r in range(3, ws.max_row + 1) if str(ws.cell(r, serial_col).value or "").strip() == serial)
    return wb, row, headers


def test_every_profile_field_roundtrips_excel_to_cache(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    p = _paths(tmp_path)
    _configure(monkeypatch, p)
    sync_excel_to_cache(
        force=True, excel_file=p["excel"], cache_file=p["cache"], state_file=p["sync_state"],
        temp_dir=p["temp"], lock_file=p["lock"], backup_dir=p["cache_backup"],
        preserve_cache_only_units=False,
    )
    cache = pd.read_csv(p["cache"], dtype=str, keep_default_na=False)
    serial = "X16B816938"
    current = cache.loc[cache["Serial Number"].eq(serial)].iloc[0]
    desired = {
        "Installation Date": "04-08-2026", "Model": "ROUNDTRIP AED MODEL",
        "Installed Phase / Month": "Phase RT", "PO Number": "PO-RT-001", "Zone": "RT Zone",
        "Block / Locations": "999", "Street Name": "Roundtrip Street", "Postal Code": "012345",
        "Level": "8", "Lift Lobby": "Lift Lobby Q",
        "Adult Pads Replacement Date": "01-08-2026", "Adult Pads Expiry Date": "01-08-2029",
        "Adult Pads Lot Number": "ADULT-ALL", "Pediatric Pads Replacement Date": "02-08-2026",
        "Pediatric Pads Expiry Date": "02-08-2029", "Pediatric Pads Lot Number": "PED-ALL",
        "Battery Replacement History": "01-01-2025; 04-08-2026", "Battery Expiry Date": "04-08-2031",
        "PM Completed Date": "04-08-2026", "Next PM Date": "04-08-2027", "Job Type": "Repair",
        "Last Done By": "Roundtrip Tester", "Service Report e-SR": "e-SR-ALL-001",
        "Repaired?": "Yes", "Remarks": "All official profile fields roundtrip test",
    }
    assert set(desired) == set(DETAIL_EDITABLE_COLUMNS)
    originals = {field: str(current.get(field, "")) for field in desired}
    result = tx.execute_unit_update(
        serial_number=serial, desired_values=desired, original_values=originals,
        user="Roundtrip Tester", session_id="all-fields", source_page="Full roundtrip test",
    )
    assert result.success, result.message
    refreshed = pd.read_csv(p["cache"], dtype=str, keep_default_na=False)
    row = refreshed.loc[refreshed["Serial Number"].eq(serial)].iloc[0]
    for field, expected in desired.items():
        assert row[field] == expected, (field, row[field], expected)
    assert row["Location"] == "Blk 999 Roundtrip Street"


def test_pm_checklist_full_commit_updates_all_linked_stores(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    p = _paths(tmp_path)
    _configure(monkeypatch, p)
    sync_excel_to_cache(
        force=True, excel_file=p["excel"], cache_file=p["cache"], state_file=p["sync_state"],
        temp_dir=p["temp"], lock_file=p["lock"], backup_dir=p["cache_backup"],
        preserve_cache_only_units=False,
    )
    pm_service.ensure_pm_storage()
    issue_service.ensure_issue_storage(p["issues"])
    df = pd.read_csv(p["cache"], dtype=str, keep_default_na=False)
    serial = "X18K075125"
    index = int(df.index[df["Serial Number"].eq(serial)][0])
    master = df.loc[index]
    pd.DataFrame([{
        "Operation ID": "", "Plan ID": "PM-2026-08", "Plan Month": "2026-08",
        "Planned Date": "04-08-2026", "Serial Number": serial, "Assigned To": "Roundtrip Tech",
        "PM Status": "Planned", "Completed Date": "", "Completed By": "",
        "Completion Record ID": "", "Completion Operation ID": "", "Is Loaner": "No",
        "Color Override": "", "Location Snapshot": master["Location"],
        "Postal Code Snapshot": master["Postal Code"], "Latitude Snapshot": master.get("Latitude", ""),
        "Longitude Snapshot": master.get("Longitude", ""), "Created At": "01-08-2026 09:00:00",
    }], columns=PM_PLAN_COLUMNS).to_csv(p["plan"], index=False, encoding="utf-8-sig")

    response = pm_service.build_response(
        service_date=date(2026, 8, 4), technician="Roundtrip Tech", service_type="PM+batt",
        customer_location="SCDF / SAL", postal_code=master["Postal Code"], lift_lobby="Z",
        loaner_unit="No", cabinet_inspection="Fail", cabinet_alarm="Pass",
        serial_number=serial, physical_condition="Pass",
        self_test_result="Pass - After installing new batteries", battery_expiry=date(2031, 8, 4),
        aed_cover="Pass", adult_pads_expiry=date(2029, 8, 4), adult_pads_lot="ADULT-RT",
        adult_pads_within_expiry="Yes", pediatric_pads_expiry=date(2028, 8, 4),
        pediatric_pads_lot="PED-RT", pediatric_pads_within_expiry="Yes",
        aed_signage="Yes", final_check="Yes", aed_location=master["Location"],
        original_serial_number=serial, service_report_id="e-SR-ROUNDTRIP",
        service_notes="Full PM roundtrip test", aed_model=master["Model"],
    )
    pm_checklist.st = SimpleNamespace(session_state={
        "audit_user": "Roundtrip Tech", "session_id": "pm-full",
        "pm_selected_model": master["Model"],
    })
    message, warnings, issue_ids = pm_checklist._commit_pm_submission(df, index, response)
    assert not warnings
    assert len(issue_ids) == 1
    assert "safely updated" in message

    wb, excel_row, headers = _excel_row(p["excel"], serial)
    ws = wb["Sheet1"]
    expected_excel = {
        "Lift Lobby": "Z", "Adult CPR-D Padz Lot Number": "ADULT-RT",
        "Children Pedi-Padz Lot Number": "PED-RT", "JOB TYPE": "PM+batt",
        "Last done by": "Roundtrip Tech", "Service Report / e-SR": "e-SR-ROUNDTRIP",
    }
    for header, expected in expected_excel.items():
        assert str(ws.cell(excel_row, headers[header]).value) == expected
    assert ws.cell(excel_row, headers["PM Completed On"]).value.date() == date(2026, 8, 4)
    assert ws.cell(excel_row, headers["Next PM Due"]).value.date() == date(2027, 8, 4)
    wb.close()

    cache = pd.read_csv(p["cache"], dtype=str, keep_default_na=False)
    master_after = cache.loc[cache["Serial Number"].eq(serial)].iloc[0]
    assert master_after["PM Completed Date"] == "04-08-2026"
    assert master_after["Next PM Date"] == "04-08-2027"
    assert master_after["Battery Replacement History"].endswith("04-08-2026")

    pm_rows = pd.read_csv(p["pm"], dtype=str, keep_default_na=False)
    saved = pm_rows.loc[pm_rows["PM Response ID"].eq(response["PM Response ID"])].iloc[0]
    assert saved["Excel Update Status"] == "UPDATED"
    assert saved["Created Issue IDs"] == issue_ids[0]
    assert saved["Linked Plan ID"] == "PM-2026-08"
    assert saved["Service Notes"] == "Full PM roundtrip test"

    plan = pd.read_csv(p["plan"], dtype=str, keep_default_na=False).iloc[0]
    assert plan["PM Status"] == "Completed"
    assert plan["Completion Record ID"] == response["PM Response ID"]
    issue = issue_service.get_issue_record(p["issues"], issue_ids[0])
    assert issue["Source"] == "PM Checklist"
    assert issue["Source Record ID"] == response["PM Response ID"]

    records = service_record_service.load_service_records(
        p["pm"], manual_service_file=p["manual"], issue_record_file=p["issues"],
        resolution_file=p["resolutions"], aed_dataframe=cache,
    )
    ledger = records.loc[records["PM Response ID"].eq(response["PM Response ID"])].iloc[0]
    assert ledger["Record Source"] == "PM Checklist"
    assert ledger["Record Match"] == "Matched"
    history = unit_profile_service.build_service_history(
        master_after, serial, pm_responses_file=p["pm"], issue_record_file=p["issues"],
        resolution_file=p["resolutions"], manual_service_file=p["manual"],
    )
    assert (history["Source"] == "PM Checklist").any()


def test_mocked_onedrive_download_update_upload_and_external_refresh(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    p = _paths(tmp_path)
    _configure(monkeypatch, p)
    remote = {"bytes": p["excel"].read_bytes(), "etag": "E1"}

    def download(*, force: bool = False):
        p["excel"].write_bytes(remote["bytes"])
        return SimpleNamespace(status="downloaded", message="downloaded", etag=remote["etag"], changed=True)

    def upload(*, expected_etag: str = ""):
        assert expected_etag == remote["etag"]
        remote["bytes"] = p["excel"].read_bytes()
        remote["etag"] = "E2"
        return SimpleNamespace(status="uploaded", message="uploaded", etag="E2")

    monkeypatch.setattr(aed_repository, "is_cloud_onedrive_enabled", lambda: True)
    monkeypatch.setattr(aed_repository, "download_workbook", download)
    monkeypatch.setattr(aed_repository, "upload_workbook", upload)
    monkeypatch.setattr(aed_repository, "EXCEL_FILE", p["excel"])
    monkeypatch.setattr(aed_repository, "AED_CACHE_FILE", p["cache"])
    monkeypatch.setattr(aed_repository, "SYNC_STATE_FILE", p["sync_state"])
    monkeypatch.setattr(aed_repository, "sync_excel_to_cache", lambda force=False, **kwargs: sync_excel_to_cache(
        force=force, excel_file=p["excel"], cache_file=p["cache"], state_file=p["sync_state"],
        temp_dir=p["temp"], lock_file=p["lock"], backup_dir=p["cache_backup"],
        preserve_cache_only_units=False,
    ))

    result = aed_repository.update_unit(
        serial_number="X18K075125", changes={"Last Done By": "Cloud Roundtrip"},
        original_values={"Last Done By": "Zihan"}, user="Cloud Tester",
        session_id="cloud", source_page="PM Checklist",
    )
    assert result.success, result.message
    remote_path = tmp_path / "remote_after.xlsx"
    remote_path.write_bytes(remote["bytes"])
    wb, row, headers = _excel_row(remote_path, "X18K075125")
    assert wb["Sheet1"].cell(row, headers["Last done by"]).value == "Cloud Roundtrip"
    wb.close()

    # External OneDrive edit must flow back into the website cache.
    wb = load_workbook(remote_path)
    ws = wb["Sheet1"]
    ws.cell(row, headers["RELATED OBJECTS"]).value = "EXTERNAL CLOUD MODEL"
    wb.save(remote_path)
    wb.close()
    remote["bytes"] = remote_path.read_bytes()
    remote["etag"] = "E3"
    refreshed = aed_repository.ensure_cache_current(force=True)
    assert refreshed.status == "synced"
    cache = pd.read_csv(p["cache"], dtype=str, keep_default_na=False)
    assert cache.loc[cache["Serial Number"].eq("X18K075125"), "Model"].iloc[0] == "EXTERNAL CLOUD MODEL"


def test_system_state_archive_roundtrips_all_operational_tables(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    root = tmp_path / "project"
    root.mkdir()
    files = []
    for name, content in {
        "pm_responses.csv": "PM Response ID,Technician\nPM-1,Tech\n",
        "pm_plan_records.csv": "Plan ID,PM Status\nPLAN-1,Completed\n",
        "issue_records.csv": "Issue ID,Status\nISS-1,Closed\n",
        "manual_service_records.csv": "Service Record ID,Status\nSRV-1,Completed\n",
        "map_unit_state.csv": "Serial Number,Status\nAED-1,Completed\n",
    }.items():
        path = root / name
        path.write_text(content, encoding="utf-8")
        files.append(path)
    photos = root / "issue_photos"
    photos.mkdir()
    (photos / "proof.jpg").write_bytes(b"proof")
    files.append(photos)
    monkeypatch.setattr(system_state_service, "PROJECT_ROOT", root)
    monkeypatch.setattr(system_state_service, "SYSTEM_STATE_PATHS", tuple(files))
    archive = system_state_service.build_archive()
    for path in files:
        if path.is_file():
            path.write_text("corrupted", encoding="utf-8")
    (photos / "proof.jpg").write_bytes(b"bad")
    system_state_service._safe_extract(archive)
    assert "PM-1,Tech" in (root / "pm_responses.csv").read_text(encoding="utf-8")
    assert "PLAN-1,Completed" in (root / "pm_plan_records.csv").read_text(encoding="utf-8")
    assert "ISS-1,Closed" in (root / "issue_records.csv").read_text(encoding="utf-8")
    assert (photos / "proof.jpg").read_bytes() == b"proof"


def _install_fake_cloud(monkeypatch: pytest.MonkeyPatch, p: dict[str, Path]):
    remote = {"bytes": p["excel"].read_bytes(), "etag": "CLOUD-E1", "uploads": 0}

    def download(*, force: bool = False):
        p["excel"].write_bytes(remote["bytes"])
        return SimpleNamespace(
            status="downloaded", message="downloaded", etag=remote["etag"],
            changed=True, source_exists=True,
        )

    def upload(*, expected_etag: str = ""):
        assert expected_etag == remote["etag"]
        remote["bytes"] = p["excel"].read_bytes()
        remote["uploads"] += 1
        remote["etag"] = f"CLOUD-E{remote['uploads'] + 1}"
        return SimpleNamespace(status="uploaded", message="uploaded", etag=remote["etag"])

    monkeypatch.setattr(aed_repository, "is_cloud_onedrive_enabled", lambda: True)
    monkeypatch.setattr(aed_repository, "download_workbook", download)
    monkeypatch.setattr(aed_repository, "upload_workbook", upload)
    monkeypatch.setattr(aed_repository, "EXCEL_FILE", p["excel"])
    monkeypatch.setattr(aed_repository, "AED_CACHE_FILE", p["cache"])
    monkeypatch.setattr(aed_repository, "SYNC_STATE_FILE", p["sync_state"])
    monkeypatch.setattr(aed_repository, "sync_excel_to_cache", lambda force=False, **kwargs: sync_excel_to_cache(
        force=force, excel_file=p["excel"], cache_file=p["cache"], state_file=p["sync_state"],
        temp_dir=p["temp"], lock_file=p["lock"], backup_dir=p["cache_backup"],
        preserve_cache_only_units=False,
    ))
    return remote


def test_pm_checklist_updates_same_mocked_onedrive_workbook(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    p = _paths(tmp_path)
    _configure(monkeypatch, p)
    remote = _install_fake_cloud(monkeypatch, p)
    aed_repository.ensure_cache_current(force=True)
    pm_service.ensure_pm_storage()
    issue_service.ensure_issue_storage(p["issues"])
    df = pd.read_csv(p["cache"], dtype=str, keep_default_na=False)
    serial = "X18K075198"
    index = int(df.index[df["Serial Number"].eq(serial)][0])
    master = df.loc[index]
    response = pm_service.build_response(
        service_date=date(2026, 8, 5), technician="Cloud PM Tech",
        service_type="Preventive Maintenance (PM)", customer_location="SCDF / SAL",
        postal_code=master["Postal Code"], lift_lobby="Y", loaner_unit="No",
        cabinet_inspection="Pass", cabinet_alarm="Pass", serial_number=serial,
        physical_condition="Pass", self_test_result="Pass", battery_expiry=date(2031, 8, 5),
        aed_cover="Pass", adult_pads_expiry=date(2029, 8, 5), adult_pads_lot="CLOUD-A",
        adult_pads_within_expiry="Yes", pediatric_pads_expiry=date(2028, 8, 5),
        pediatric_pads_lot="CLOUD-P", pediatric_pads_within_expiry="Yes",
        aed_signage="Yes", final_check="Yes", aed_location=master["Location"],
        original_serial_number=serial, service_report_id="e-SR-CLOUD-PM",
        service_notes="Cloud PM roundtrip", aed_model=master["Model"],
    )
    pm_checklist.st = SimpleNamespace(session_state={
        "audit_user": "Cloud PM Tech", "session_id": "pm-cloud",
        "pm_selected_model": master["Model"],
    })
    message, warnings, issue_ids = pm_checklist._commit_pm_submission(df, index, response)
    assert not warnings and not issue_ids
    assert remote["uploads"] == 1
    assert "OneDrive" in message or "IB List" in message
    remote_file = tmp_path / "remote_pm.xlsx"
    remote_file.write_bytes(remote["bytes"])
    wb, row, headers = _excel_row(remote_file, serial)
    ws = wb["Sheet1"]
    assert ws.cell(row, headers["Last done by"]).value == "Cloud PM Tech"
    assert ws.cell(row, headers["Service Report / e-SR"]).value == "e-SR-CLOUD-PM"
    assert ws.cell(row, headers["PM Completed On"]).value.date() == date(2026, 8, 5)
    wb.close()
    pm_rows = pd.read_csv(p["pm"], dtype=str, keep_default_na=False)
    assert pm_rows.loc[pm_rows["PM Response ID"].eq(response["PM Response ID"]), "Excel Update Status"].iloc[0] == "UPDATED"


def test_manual_service_updates_excel_plan_service_records_and_profile(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    p = _paths(tmp_path)
    _configure(monkeypatch, p)
    sync_excel_to_cache(
        force=True, excel_file=p["excel"], cache_file=p["cache"], state_file=p["sync_state"],
        temp_dir=p["temp"], lock_file=p["lock"], backup_dir=p["cache_backup"],
        preserve_cache_only_units=False,
    )
    cache = pd.read_csv(p["cache"], dtype=str, keep_default_na=False)
    serial = "X16A810545"
    master = cache.loc[cache["Serial Number"].eq(serial)].iloc[0]
    pd.DataFrame([{
        "Operation ID": "", "Plan ID": "PM-2026-08", "Plan Month": "2026-08",
        "Planned Date": "06-08-2026", "Serial Number": serial, "Assigned To": "Manual Tech",
        "PM Status": "Planned", "Completed Date": "", "Completed By": "",
        "Completion Record ID": "", "Completion Operation ID": "", "Is Loaner": "No",
        "Color Override": "", "Location Snapshot": master["Location"],
        "Postal Code Snapshot": master["Postal Code"], "Latitude Snapshot": master.get("Latitude", ""),
        "Longitude Snapshot": master.get("Longitude", ""), "Created At": "01-08-2026 09:00:00",
    }], columns=PM_PLAN_COLUMNS).to_csv(p["plan"], index=False, encoding="utf-8-sig")
    plan = unit_profile_service.build_manual_service_update_plan(
        master, service_date=date(2026, 8, 6), service_type="PM+batt",
        technician="Manual Tech", reference="e-SR-MANUAL", status="Completed",
        update_latest=True, update_pm_dates=True, interval_months=12,
    )
    result = aed_repository.update_unit(
        serial_number=serial, changes=plan["changes"], original_values=plan["originals"],
        user="Manual Tech", session_id="manual", source_page="AED Management Add Service Record",
    )
    assert result.success, result.message
    record_id = "SRV-MANUAL-ROUNDTRIP"
    linked = pm_service.complete_matching_pm_plan(
        serial, "06-08-2026", operation_id=result.operation_id, response_id=record_id,
        completed_by="Manual Tech", plan_file=p["plan"],
    )
    saved = unit_profile_service.append_manual_service_record({
        "Service Record ID": record_id, "Created At": "06-08-2026 12:00:00",
        "Created By": "Manual Tech", "AED Serial Number": serial,
        "AED Model": master["Model"], "AED Location": master["Location"],
        "Postal Code": master["Postal Code"], "Lift Lobby": master["Lift Lobby"],
        "Service Date": "06-08-2026", "Service Type": "PM+batt", "Technician": "Manual Tech",
        "Reference": "e-SR-MANUAL", "Status": "Completed", "Details": "Manual service roundtrip",
        "Master Data Updated": "Yes", "PM Dates Updated": "Yes", "Battery Replaced": "Yes",
        "Battery History Updated": "Yes", "PM Interval Months Used": "12",
        "Linked Plan ID": linked, "Master Operation ID": result.operation_id, "Source": "Unit Profile",
    }, path=p["manual"])
    assert saved["Service Record ID"] == record_id
    refreshed = pd.read_csv(p["cache"], dtype=str, keep_default_na=False)
    after = refreshed.loc[refreshed["Serial Number"].eq(serial)].iloc[0]
    assert after["PM Completed Date"] == "06-08-2026"
    assert after["Next PM Date"] == "06-08-2027"
    assert after["Service Report e-SR"] == "e-SR-MANUAL"
    assert after["Battery Replacement History"].endswith("06-08-2026")
    assert pd.read_csv(p["plan"], dtype=str, keep_default_na=False).iloc[0]["PM Status"] == "Completed"
    records = service_record_service.load_service_records(
        p["pm"], manual_service_file=p["manual"], issue_record_file=p["issues"],
        resolution_file=p["resolutions"], aed_dataframe=refreshed,
    )
    assert (records["Service Record ID"] == record_id).any()
    history = unit_profile_service.build_service_history(
        after, serial, pm_responses_file=p["pm"], issue_record_file=p["issues"],
        resolution_file=p["resolutions"], manual_service_file=p["manual"],
    )
    assert ((history["Source"] == "Unit Profile") & history["Reference"].eq("e-SR-MANUAL")).any()


def test_loaner_pm_is_saved_but_does_not_change_official_excel(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    p = _paths(tmp_path)
    _configure(monkeypatch, p)
    sync_excel_to_cache(
        force=True, excel_file=p["excel"], cache_file=p["cache"], state_file=p["sync_state"],
        temp_dir=p["temp"], lock_file=p["lock"], backup_dir=p["cache_backup"],
        preserve_cache_only_units=False,
    )
    pm_service.ensure_pm_storage()
    issue_service.ensure_issue_storage(p["issues"])
    before = p["excel"].read_bytes()
    df = pd.read_csv(p["cache"], dtype=str, keep_default_na=False)
    serial = "LOANER-TEST-001"
    # The selected master row is only used for display snapshots in the loaner branch.
    index = int(df.index[0])
    response = pm_service.build_response(
        service_date=date(2026, 8, 7), technician="Loaner Tech",
        service_type="Preventive Maintenance (PM)", customer_location="Other",
        postal_code="123456", lift_lobby="A", loaner_unit="Yes",
        cabinet_inspection="Pass", cabinet_alarm="Pass", serial_number=serial,
        physical_condition="Pass", self_test_result="Pass", battery_expiry=None,
        aed_cover="Pass", adult_pads_expiry=None, adult_pads_lot="",
        adult_pads_within_expiry="Yes", pediatric_pads_expiry=None,
        pediatric_pads_lot="", pediatric_pads_within_expiry="Yes",
        aed_signage="Yes", final_check="Yes", aed_location="Temporary Loaner",
        original_serial_number=serial, service_report_id="LOANER-REF",
        service_notes="Loaner PM", aed_model="AED Plus",
    )
    pm_checklist.st = SimpleNamespace(session_state={
        "audit_user": "Loaner Tech", "session_id": "loaner",
        "pm_selected_model": "AED Plus",
    })
    pm_checklist._commit_pm_submission(df, index, response)
    assert p["excel"].read_bytes() == before
    saved = pd.read_csv(p["pm"], dtype=str, keep_default_na=False).iloc[0]
    assert saved["Loaner Unit"] == "Yes"
    assert saved["Excel Update Status"] == "NOT_REQUIRED_LOANER"
    assert saved["Master Data Updated"] == "No"


def test_pm_operational_records_upload_to_separate_onedrive_state_archive(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    import zipfile

    root = tmp_path / "project"
    root.mkdir()
    pm_file = root / "pm_responses.csv"
    plan_file = root / "pm_plan_records.csv"
    issue_file = root / "issue_records.csv"
    pm_file.write_text("PM Response ID,Technician\nPM-CLOUD-1,Tech\n", encoding="utf-8")
    plan_file.write_text("Plan ID,PM Status\nPLAN-CLOUD-1,Completed\n", encoding="utf-8")
    issue_file.write_text("Issue ID,Status\nISS-CLOUD-1,Reported\n", encoding="utf-8")
    state_file = root / "data" / "system_state_sync.json"
    pending_dir = root / "backups" / "pending"
    captured: dict[str, object] = {}

    monkeypatch.setattr(system_state_service, "ONEDRIVE_CLOUD_ENABLED", True)
    monkeypatch.setattr(system_state_service, "PROJECT_ROOT", root)
    monkeypatch.setattr(
        system_state_service, "SYSTEM_STATE_PATHS", (pm_file, plan_file, issue_file)
    )
    monkeypatch.setattr(system_state_service, "SYSTEM_STATE_SYNC_FILE", state_file)
    monkeypatch.setattr(system_state_service, "SYSTEM_STATE_PENDING_DIR", pending_dir)
    monkeypatch.setattr(system_state_service, "get_metadata", lambda path, missing_ok=True: None)

    def fake_upload(path, content, *, content_type, expected_etag=""):
        captured["path"] = path
        captured["content"] = content
        captured["content_type"] = content_type
        return SimpleNamespace(etag="STATE-E1")

    monkeypatch.setattr(system_state_service, "upload_bytes", fake_upload)
    result = system_state_service.sync_system_state()
    assert result.uploaded
    assert captured["content_type"] == "application/zip"
    with zipfile.ZipFile(io.BytesIO(captured["content"]), "r") as archive:
        assert set(archive.namelist()) == {
            "pm_responses.csv", "pm_plan_records.csv", "issue_records.csv"
        }
        assert "PM-CLOUD-1,Tech" in archive.read("pm_responses.csv").decode("utf-8")
